#!/usr/bin/env python3
"""End-to-end relist demo: real Ozon data -> restyled images -> verified xlsx.

Two modes:

  --mock   Fully offline. Fake Ozon data, a fake imagegen (returns a tiny fixed
           PNG) and a fake host (deterministic fake URLs). NO network, NO spend.
           This is what CI/tests run.

  (real)   Pull N products from the Ozon Seller API (OzonClient.from_env),
           inject their name/description/images into the core sample template,
           run the core `process` with the real relist transform
           (gptimage_edit + get_host), verify, then independently fetch each NEW
           image URL and assert it is a 200 image.

Pipeline (both modes):
  1. gen-sample           -> the core template workbook
  2. mapper + fields.yaml  -> locate title/listing/image columns + data start
  3. inject real product data into the first N data rows -> real_input.xlsx
  4. core process(--transform relist) -> real_output.xlsx
  5. core verify           -> prove ONLY title/listing/image cells changed
  6. fetch each new image URL -> assert HTTP 200 image (real mode)

Usage:
  # offline (no network, no keys, no spend):
  .venv/bin/python scripts/demo_real.py --mock

  # real (needs env vars; see .env.example):
  export OZON_CLIENT_ID=... OZON_API_KEY=...
  export OZON_RELIST_IMAGE_API_KEY=...
  # optional: OZON_RELIST_IMAGE_API_BASE, OZON_RELIST_IMAGE_MODEL,
  #           OZON_RELIST_HOST, OZON_RELIST_CACHE, OZON_RELIST_PROMPT
  .venv/bin/python scripts/demo_real.py --products 3 --out-dir ./relist_demo
"""

from __future__ import annotations

import argparse
import os
import sys

# Make the package importable when run as a bare script (editable install also
# works; this is belt-and-suspenders for `python scripts/demo_real.py`).
_HERE = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.dirname(_HERE)
sys.path.insert(0, os.path.join(_ROOT, "src"))

from openpyxl import load_workbook  # noqa: E402

from ozon_excel_core import process, verify  # noqa: E402
from ozon_excel_core.config import load_config  # noqa: E402
from ozon_excel_core.header import detect_header  # noqa: E402
from ozon_excel_core.mapper import resolve_columns  # noqa: E402
from ozon_excel_core.sample import gen_sample  # noqa: E402
from ozon_excel_core.transforms.relist import RelistTransform  # noqa: E402

CONFIG_PATH = os.path.join(_ROOT, "fields.example.yaml")
MAIN_SHEET = "Шаблон для поставщика"

# A 1x1 transparent PNG — the fixed fake "generated" image used in --mock mode.
_TINY_PNG = bytes(
    [
        0x89, 0x50, 0x4E, 0x47, 0x0D, 0x0A, 0x1A, 0x0A, 0x00, 0x00, 0x00, 0x0D,
        0x49, 0x48, 0x44, 0x52, 0x00, 0x00, 0x00, 0x01, 0x00, 0x00, 0x00, 0x01,
        0x08, 0x06, 0x00, 0x00, 0x00, 0x1F, 0x15, 0xC4, 0x89, 0x00, 0x00, 0x00,
        0x0D, 0x49, 0x44, 0x41, 0x54, 0x78, 0x9C, 0x62, 0x00, 0x01, 0x00, 0x00,
        0x05, 0x00, 0x01, 0x0D, 0x0A, 0x2D, 0xB4, 0x00, 0x00, 0x00, 0x00, 0x49,
        0x45, 0x4E, 0x44, 0xAE, 0x42, 0x60, 0x82,
    ]
)


# --------------------------------------------------------------------------- #
# Fakes for --mock mode (no network, no spend)
# --------------------------------------------------------------------------- #
class FakeImageGen:
    """Stand-in for gptimage_edit: returns a fixed tiny PNG, counts calls."""

    def __init__(self):
        self.calls = 0

    def __call__(self, image_bytes, prompt, **kwargs):
        self.calls += 1
        return _TINY_PNG


class FakeHost:
    """Stand-in for a real host: returns a deterministic fake URL per filename,
    and records the bytes so we can serve them back to the demo's URL fetch."""

    name = "fake"

    def __init__(self, base="https://fake-host.test/relist"):
        self.base = base.rstrip("/")
        self.store = {}  # url -> bytes

    def put(self, filename, data):
        url = f"{self.base}/{filename}"
        self.store[url] = data
        return url


def _fake_products(n):
    """Deterministic fake Ozon products (name, description, image urls)."""
    catalog = [
        {
            "name": "  чайник  электрический  стеклянный  ",
            "description": "Объём 1.7 л.\r\nМощность 2200 Вт.\n\n\nГарантия 12 месяцев.",
            "primary_image": "https://img.ozon.test/1001/main.jpg",
            "images": [
                "https://img.ozon.test/1001/a1.jpg",
                "https://img.ozon.test/1001/p2.jpg",
            ],
        },
        {
            "name": "Кружка керамическая 350 мл, синяя",
            "description": "Подходит для микроволновой печи.",
            "primary_image": "https://img.ozon.test/1002/main.jpg",
            "images": [
                "https://img.ozon.test/1002/a1.jpg",
                "https://img.ozon.test/1002/p2.jpg",
            ],
        },
        {
            "name": "Нож кухонный шеф 20 см",
            "description": "Нержавеющая сталь. Эргономичная ручка.",
            "primary_image": "https://img.ozon.test/1003/main.jpg",
            "images": [
                "https://img.ozon.test/1003/a1.jpg",
                "https://img.ozon.test/1003/p2.jpg",
            ],
        },
        {
            "name": "Полотенце махровое 70x140, белое",
            "description": "Хлопок 100%. Плотность 450 г/м².",
            "primary_image": "https://img.ozon.test/1004/main.jpg",
            "images": [
                "https://img.ozon.test/1004/a1.jpg",
                "https://img.ozon.test/1004/p2.jpg",
            ],
        },
        {
            "name": "Лампа настольная LED, чёрная",
            "description": "3 режима яркости. USB-питание.",
            "primary_image": "https://img.ozon.test/1005/main.jpg",
            "images": [
                "https://img.ozon.test/1005/a1.jpg",
                "https://img.ozon.test/1005/p2.jpg",
            ],
        },
    ]
    out = []
    i = 0
    while len(out) < n:
        out.append(catalog[i % len(catalog)])
        i += 1
    return out


# --------------------------------------------------------------------------- #
# Real Ozon fetch
# --------------------------------------------------------------------------- #
def _real_products(n):
    """Fetch N real products (name, description, primary/other images)."""
    from ozon_excel_core.ozon_api import OzonClient

    client = OzonClient.from_env()
    items = client.list_products(limit=n)[:n]
    product_ids = [it.get("product_id") for it in items if it.get("product_id")]
    offer_by_pid = {it.get("product_id"): it.get("offer_id") for it in items}
    infos = client.product_info(product_ids)

    out = []
    for info in infos[:n]:
        pid = info.get("id") or info.get("product_id")
        offer_id = offer_by_pid.get(pid)
        try:
            desc = client.product_description(
                offer_id=offer_id, product_id=None if offer_id else pid
            )
        except Exception as exc:  # description is best-effort
            desc = f"(description unavailable: {exc})"
        out.append(
            {
                "name": info.get("name") or "",
                "description": desc,
                "primary_image": info.get("primary_image") or "",
                "images": list(info.get("images") or []),
            }
        )
    return out


# --------------------------------------------------------------------------- #
# Injection: write real product data into the mapped target columns
# --------------------------------------------------------------------------- #
def _inject_products(in_xlsx, out_xlsx, config, products):
    """Use the mapper to find the title/listing/image columns + data start, then
    overwrite the first len(products) data rows with the real product data.

    Thin adapter over ozon_excel_core.inject.inject_rows: it maps the demo's
    product dicts (name/description/primary_image/images) onto the package's
    canonical row shape (title/listing/primary_image/additional_images) and
    applies the demo's per-product image cap before delegating. Returns a dict
    describing where each piece landed (for the report)."""
    from ozon_excel_core.inject import inject_rows

    _img_cap = int(os.environ.get("OZON_RELIST_DEMO_IMG_CAP", "2"))
    rows = [
        {
            "title": str(product["name"]),
            "listing": str(product.get("description") or ""),
            "primary_image": product.get("primary_image"),
            "additional_images": list(product.get("images") or [])[:_img_cap],
        }
        for product in products
    ]
    placed = inject_rows(in_xlsx, out_xlsx, config, rows)
    # Preserve the demo report's historical "name" key (inject_rows reports
    # "title"); the rest of the placement dict is identical.
    placed["rows"] = [
        {"row": r["row"], "name": r["title"]} for r in placed["rows"]
    ]
    return placed


# --------------------------------------------------------------------------- #
# Verify + fetch new image URLs
# --------------------------------------------------------------------------- #
def _collect_new_image_urls(out_xlsx, config, n_rows):
    """Read back the image target cells from the processed output and return the
    list of (coordinate, url) for the first n_rows data rows."""
    wb = load_workbook(str(out_xlsx))
    ws = wb[MAIN_SHEET]
    hb = detect_header(ws, config)
    mapped = resolve_columns(ws, config, hb)
    image_cols = [mc.col_index for mc in (mapped.images_main + mapped.images_additional)]
    data_start = hb.data_start_row

    found = []
    import re as _re

    url_re = _re.compile(r"https?://\S+")
    for i in range(n_rows):
        r = data_start + i
        for col in image_cols:
            cell = ws.cell(row=r, column=col)
            val = cell.value
            if not isinstance(val, str):
                continue
            for m in url_re.findall(val):
                found.append((cell.coordinate, m.rstrip('"),')))
    wb.close()
    return found


def _fetch_status(url, fake_host=None):
    """Return (ok, detail). In mock mode serve from the fake host's store; in
    real mode do a real GET and assert it is a 200 image."""
    if fake_host is not None:
        data = fake_host.store.get(url)
        if data is None:
            return False, "not in fake host store"
        is_png = data[:8] == b"\x89PNG\r\n\x1a\n"
        return True, f"200 fake-image ({len(data)} bytes, png={is_png})"
    import requests

    try:
        resp = requests.get(url, timeout=60)
    except Exception as exc:
        return False, f"fetch error: {exc}"
    ctype = resp.headers.get("Content-Type", "")
    ok = resp.status_code == 200 and (
        ctype.startswith("image/") or resp.content[:8] == b"\x89PNG\r\n\x1a\n"
    )
    return ok, f"HTTP {resp.status_code} {ctype} ({len(resp.content)} bytes)"


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def run(mock: bool, products_n: int, out_dir: str) -> int:
    os.makedirs(out_dir, exist_ok=True)
    config = load_config(CONFIG_PATH)

    template = os.path.join(out_dir, "template.xlsx")
    real_input = os.path.join(out_dir, "real_input.xlsx")
    real_output = os.path.join(out_dir, "real_output.xlsx")

    print(f"=== relist demo ({'MOCK / offline' if mock else 'REAL'}) ===")
    print(f"products: {products_n}   out-dir: {os.path.abspath(out_dir)}")

    # 1. template
    gen_sample(template)

    # 2. product data
    if mock:
        products = _fake_products(products_n)
    else:
        products = _real_products(products_n)
    print(f"fetched {len(products)} products")

    # 3. inject -> real_input.xlsx
    placed = _inject_products(template, real_input, config, products)
    print(f"injected into rows starting at {placed['data_start_row']}:")
    print(
        f"  title={placed['title_col']}  listing={placed['listing_col']}  "
        f"main_image={placed['main_image_col']}  "
        f"additional_images={placed['additional_image_cols']}"
    )
    for row in placed["rows"]:
        print(f"    row {row['row']}: {row['name']}")

    # 4. build the relist transform (injected fakes in mock mode)
    fake_host = None
    if mock:
        fake_host = FakeHost()
        transform = RelistTransform(
            imagegen_fn=FakeImageGen(),
            host=fake_host,
            image_api_base="https://fake.test",
            image_api_key="fake-key",
            image_model="gpt-image-2-vip",
            cache_dir=os.path.join(out_dir, "cache"),
        )

        # Offline source-image download: monkeypatch requests.get used by the
        # transform so the *source* ozon.test urls resolve without network.
        import ozon_excel_core.transforms.relist as relist_mod

        class _Resp:
            status_code = 200
            content = _TINY_PNG

        _orig_get = relist_mod.requests.get
        relist_mod.requests.get = lambda url, **kw: _Resp()
    else:
        from ozon_excel_core.transforms.relist import make_relist

        transform = make_relist()

    # 4b. core process -> real_output.xlsx
    try:
        result = process(real_input, real_output, config, transform)
    finally:
        if mock:
            relist_mod.requests.get = _orig_get

    print(f"process: {result.total_changed()} cells changed "
          f"({dict(result.changed_by_role)})")

    # 5. verify (independent re-derivation)
    report = verify(real_input, real_output, config)
    roles = sorted({d.role for d in report.expected_changes})
    print(f"verify: ok={report.ok}  expected={len(report.expected_changes)} "
          f"unexpected={len(report.unexpected_changes)}  roles={roles}")
    print(f"  frozen cells compared: {report.summary.get('frozen_cells_compared')}")
    if not report.ok:
        print("VERIFY FAILED — unexpected changes:")
        for d in report.unexpected_changes:
            print(f"  {d.sheet}!{d.coordinate} [{d.role}/{d.field}]")
        return 1
    # Confirm only content roles changed.
    if not set(roles) <= {"title", "listing", "image"}:
        print(f"VERIFY: unexpected non-content roles changed: {roles}")
        return 1

    # 6. fetch each NEW image url and assert it is a 200 image
    new_urls = _collect_new_image_urls(real_output, config, len(products))
    print(f"new image URLs ({len(new_urls)}):")
    all_ok = True
    for coord, url in new_urls:
        ok, detail = _fetch_status(url, fake_host=fake_host)
        all_ok = all_ok and ok
        print(f"  {coord}: {url}  -> {'OK' if ok else 'FAIL'} [{detail}]")

    print()
    print(f"RESULT: {'OK' if (report.ok and all_ok) else 'FAIL'} — "
          f"only title/listing/image cells changed; "
          f"{len(new_urls)} new image url(s) "
          f"{'all reachable' if all_ok else 'NOT all reachable'}.")
    return 0 if (report.ok and all_ok) else 1


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="End-to-end relist demo.")
    ap.add_argument("--mock", action="store_true",
                    help="run fully offline with fakes (no network, no spend)")
    ap.add_argument("--products", type=int, default=3,
                    help="number of products to relist (default 3)")
    ap.add_argument("--out-dir", default="relist_demo",
                    help="output directory for the demo artifacts")
    args = ap.parse_args(argv)
    return run(args.mock, args.products, args.out_dir)


if __name__ == "__main__":
    raise SystemExit(main())
