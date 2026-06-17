#!/usr/bin/env python3
"""push-ozon: take a relisted output .xlsx and push the changes back to Ozon.

This is the write-side counterpart to the surgical core. The core only ever
*reads* the .xlsx (this script never mutates the workbook); the relisted file is
treated as the post-relist source of truth and its image-column URLs / title
cell are pushed to the Ozon Seller API.

Two writes are supported:

  IMAGES (default, safe)  POST /v1/product/pictures/import with the row's image
                          URLs in order (images[0] becomes the new primary).
                          --keep-existing appends the product's current images
                          after the new ones.

  TITLE (opt-in)          --push-title fetches /v4/product/info/attributes,
                          replaces ONLY the name, and re-submits the product via
                          /v1/product/import (then polls /v1/product/import/info).
                          This re-submits the whole product; it is OFF by default
                          and prints a WARNING. If the attributes response is
                          missing fields needed to round-trip safely, the title
                          push is SKIPPED for that product rather than risk
                          corrupting it.

SAFETY: --dry-run is the DEFAULT. It prints a per-product PLAN and makes ZERO
API writes. --apply performs the writes. --limit N (default 1) caps how many
products are touched. After --apply the touched products are re-fetched and the
new primary_image is confirmed.

Usage:
  # plan only (no creds needed for the read; resolution still calls the API):
  .venv/bin/python scripts/push_ozon.py --in relisted.xlsx --config fields.example.yaml

  # actually push images for the first product:
  export OZON_CLIENT_ID=... OZON_API_KEY=...
  .venv/bin/python scripts/push_ozon.py --in relisted.xlsx --config fields.example.yaml --apply
"""

from __future__ import annotations

import argparse
import os
import re
import sys

# Make the package importable when run as a bare script.
_HERE = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.dirname(_HERE)
sys.path.insert(0, os.path.join(_ROOT, "src"))

from openpyxl import load_workbook  # noqa: E402
from openpyxl.utils import column_index_from_string, get_column_letter  # noqa: E402

from ozon_excel_core import images as images_mod  # noqa: E402
from ozon_excel_core.config import load_config  # noqa: E402
from ozon_excel_core.header import detect_header  # noqa: E402
from ozon_excel_core.mapper import resolve_columns  # noqa: E402

# Title attribute id on Ozon import items. Name is its own top-level "name"
# field on the import item, so we set it directly and preserve attributes as-is.
_NAME_ATTRIBUTE_ID = 4180


# --------------------------------------------------------------------------- #
# Row reading (READ-ONLY — never mutates the workbook)
# --------------------------------------------------------------------------- #
class PushRow:
    """One relisted product row: offer_id, ordered image urls, title cell."""

    def __init__(self, row, offer_id, image_urls, title, image_coords):
        self.row = row
        self.offer_id = offer_id
        self.image_urls = image_urls  # list[str], in column order; [0] = primary
        self.title = title  # str | None
        self.image_coords = image_coords  # list[str] for the report


def _resolve_offer_col(ws, config, mapped, offer_col_arg):
    """Decide which column holds the offer_id.

    Precedence:
      1. --offer-col as a column LETTER (e.g. "A")
      2. --offer-col as a header KEYWORD matched against the header row
      3. the mapper's key_column (Артикул / SKU / Offer ID)
    """
    match_row = mapped.header_block.match_row
    max_col = ws.max_column or 0
    if offer_col_arg:
        arg = str(offer_col_arg).strip()
        # 1. column LETTER: ASCII A-Z only, resolving to a column that exists in
        #    the sheet. (A short alpha keyword like "SKU" would resolve to a
        #    nonexistent far-right column, so the in-range check disambiguates it
        #    from a header keyword.)
        if re.fullmatch(r"[A-Za-z]{1,3}", arg):
            try:
                idx = column_index_from_string(arg.upper())
            except ValueError:
                idx = None
            if idx is not None and idx <= max_col:
                return idx
        # 2. header KEYWORD (case-insensitive, normalized like the core)
        normalize = config.normalizer()
        needle = normalize(arg)
        for col in range(1, max_col + 1):
            v = ws.cell(row=match_row, column=col).value
            if v is None:
                continue
            if needle and needle in normalize(v):
                return col
        raise SystemExit(
            f"--offer-col {offer_col_arg!r} did not match a column letter or a "
            f"header keyword on row {match_row}."
        )
    # default: the mapper's key column
    if mapped.key_column is not None:
        return mapped.key_column.col_index
    raise SystemExit(
        "No offer_id column found. Pass --offer-col <letter or header keyword> "
        "(the config has no key_column to fall back on)."
    )


def _image_columns(mapped) -> list:
    """Image target columns in order: main first, then additional, as resolved."""
    return [mc.col_index for mc in (mapped.images_main + mapped.images_additional)]


def _field_for(config, col_index, mapped):
    """The FieldSpec to use when parsing an image cell (gives read_delimiters)."""
    for mc, spec in zip(mapped.images_main, config.images_main):
        if mc.col_index == col_index:
            return spec
    for mc, spec in zip(mapped.images_additional, config.images_additional):
        if mc.col_index == col_index:
            return spec
    return None


def read_rows(in_xlsx, config, *, sheet=None, offer_col_arg=None) -> list:
    """Read the relisted workbook into PushRow objects. READ-ONLY."""
    wb = load_workbook(str(in_xlsx), data_only=False, keep_links=True)
    try:
        if sheet is not None:
            ws = wb[sheet]
        else:
            ws = None
            for cand in wb.worksheets:
                if getattr(cand, "sheet_state", "visible") != "visible":
                    continue
                if config.matches_processed_sheet(cand.title):
                    ws = cand
                    break
            if ws is None:
                ws = wb.worksheets[0]

        hb = detect_header(ws, config)
        mapped = resolve_columns(ws, config, hb)
        offer_col = _resolve_offer_col(ws, config, mapped, offer_col_arg)
        image_cols = _image_columns(mapped)
        title_col = mapped.title.col_index if mapped.title else None
        data_start = hb.data_start_row

        rows = []
        for r in range(data_start, (ws.max_row or data_start - 1) + 1):
            offer_v = ws.cell(row=r, column=offer_col).value
            offer_id = None if offer_v is None else str(offer_v).strip()
            if not offer_id:
                continue  # skip blank rows

            urls = []
            coords = []
            for col in image_cols:
                cell = ws.cell(row=r, column=col)
                spec = _field_for(config, col, mapped)
                ic = images_mod.parse(cell, field_spec=spec, ws=ws)
                for u in ic.urls:
                    if u and u.strip():
                        urls.append(u.strip())
                        coords.append(cell.coordinate)

            title = None
            if title_col is not None:
                tv = ws.cell(row=r, column=title_col).value
                title = None if tv is None else str(tv)

            rows.append(PushRow(r, offer_id, urls, title, coords))
        return rows, _layout(mapped, offer_col, image_cols, title_col)
    finally:
        wb.close()


def _layout(mapped, offer_col, image_cols, title_col):
    return {
        "offer_col": get_column_letter(offer_col),
        "image_cols": [get_column_letter(c) for c in image_cols],
        "title_col": get_column_letter(title_col) if title_col else None,
        "data_start_row": mapped.header_block.data_start_row,
    }


# --------------------------------------------------------------------------- #
# product_id resolution
# --------------------------------------------------------------------------- #
def resolve_product_ids(client, offer_ids) -> dict:
    """Batch-resolve offer_id -> {product_id, name, primary_image, images} via
    /v3/product/info/list (one call)."""
    if not offer_ids:
        return {}
    items = client.product_info_by_offer(list(offer_ids))
    out = {}
    for it in items:
        oid = it.get("offer_id")
        if oid is None:
            continue
        out[str(oid)] = {
            "product_id": it.get("id") or it.get("product_id"),
            "name": it.get("name") or "",
            "primary_image": it.get("primary_image") or "",
            "images": list(it.get("images") or []),
        }
    return out


# --------------------------------------------------------------------------- #
# Push planning + execution
# --------------------------------------------------------------------------- #
def build_image_list(row_urls, existing_images, keep_existing) -> list:
    """The image list to send. row_urls first (in order); when keep_existing,
    append the product's current images that aren't already present."""
    images = list(row_urls)
    if keep_existing:
        seen = set(images)
        for u in existing_images:
            if u and u not in seen:
                images.append(u)
                seen.add(u)
    return images


def build_title_import_item(attrs, new_title):
    """Build a /v1/product/import item that changes ONLY the name and preserves
    every other attribute/field. Returns (item, None) on success, or
    (None, reason) when the attributes are too thin to round-trip safely."""
    if not attrs:
        return None, "no attributes returned"

    # Fields the import item must echo back to avoid wiping product data. If any
    # are missing we refuse rather than risk corrupting the product.
    required = ["offer_id", "category_id", "attributes"]
    # Ozon has migrated category_id -> description_category_id/type_id over time;
    # accept either shape.
    has_category = (
        attrs.get("category_id")
        or attrs.get("description_category_id")
        or attrs.get("type_id")
    )
    missing = [f for f in ("offer_id", "attributes") if attrs.get(f) in (None, "", [])]
    if not has_category:
        missing.append("category_id/description_category_id+type_id")
    if missing:
        return None, f"attributes missing {missing}; skipping to avoid corruption"

    item = dict(attrs)  # preserve every field Ozon gave us
    # Set the human-visible name. Ozon import items carry the name both as a
    # top-level "name" and (for many categories) as attribute id 4180; update
    # both so the change is consistent, touching nothing else.
    item["name"] = new_title
    new_attrs = []
    for a in attrs.get("attributes") or []:
        if a.get("id") == _NAME_ATTRIBUTE_ID or a.get("attribute_id") == _NAME_ATTRIBUTE_ID:
            a = dict(a)
            a["values"] = [{"value": new_title}]
        new_attrs.append(a)
    item["attributes"] = new_attrs
    return item, None


def plan_product(client, row, resolved, *, keep_existing, push_title):
    """Compute the per-product plan dict (no writes)."""
    info = resolved.get(row.offer_id)
    plan = {
        "offer_id": row.offer_id,
        "row": row.row,
        "product_id": info["product_id"] if info else None,
        "resolved": info is not None,
        "image_urls": list(row.image_urls),
        "primary": row.image_urls[0] if row.image_urls else None,
        "images_to_set": [],
        "keep_existing": keep_existing,
        "title": row.title,
        "title_push": False,
        "title_skip_reason": None,
        "current_primary": info["primary_image"] if info else None,
    }
    if info is None:
        plan["skip_reason"] = "offer_id not found via /v3/product/info/list"
        return plan

    plan["images_to_set"] = build_image_list(
        row.image_urls, info["images"], keep_existing
    )

    if push_title and row.title is not None:
        attrs = client.product_attributes(info["product_id"])
        item, reason = build_title_import_item(attrs, row.title)
        if item is None:
            plan["title_skip_reason"] = reason
        else:
            plan["title_push"] = True
            plan["_title_item"] = item
    return plan


def apply_product(client, plan, *, poll, no_images=False):
    """Execute the writes described by a plan. Returns a result dict."""
    res = {"offer_id": plan["offer_id"], "product_id": plan["product_id"]}
    pid = plan["product_id"]

    # 1. images (skipped entirely with --no-images, e.g. for a title-only push)
    if plan["images_to_set"] and not no_images:
        result = client.pictures_import(pid, plan["images_to_set"])
        pics = result.get("pictures") or []
        res["pictures"] = [
            {
                "url": p.get("url"),
                "state": p.get("state"),
                "is_primary": p.get("is_primary"),
            }
            for p in pics
        ]
    else:
        res["pictures"] = []

    # 2. title (optional)
    if plan.get("title_push"):
        task_id = client.product_import([plan["_title_item"]])
        res["title_task_id"] = task_id
        statuses = []
        for _ in range(max(1, poll)):
            info = client.product_import_info(task_id)
            statuses.append(info.get("status") or info)
        res["title_status"] = statuses
    return res


def confirm_product(client, plan) -> dict:
    """Re-fetch a touched product and report whether primary_image reflects the
    pushed primary image."""
    items = client.product_info_by_offer([plan["offer_id"]])
    info = items[0] if items else {}
    new_primary = info.get("primary_image") or ""
    expected = plan["primary"]
    return {
        "offer_id": plan["offer_id"],
        "new_primary": new_primary,
        "expected_primary": expected,
        "match": bool(expected) and new_primary == expected,
    }


# --------------------------------------------------------------------------- #
# Reporting
# --------------------------------------------------------------------------- #
def print_plan(plan, out=print):
    out(f"--- offer_id {plan['offer_id']} (row {plan['row']}) ---")
    if not plan["resolved"]:
        out(f"  SKIP: {plan.get('skip_reason')}")
        return
    out(f"  product_id : {plan['product_id']}")
    out(f"  current primary_image: {plan['current_primary'] or '(none)'}")
    imgs = plan["images_to_set"]
    if imgs:
        out(f"  images to set ({len(imgs)})"
            f"{' [appending existing]' if plan['keep_existing'] else ''}:")
        for i, u in enumerate(imgs):
            tag = "PRIMARY" if i == 0 else f"#{i + 1}"
            out(f"    [{tag}] {u}")
    else:
        out("  images to set: (none — no image URLs in this row)")
    if plan["title_push"]:
        out(f"  title change QUEUED -> {plan['title']!r}")
    elif plan["title_skip_reason"]:
        out(f"  title change SKIPPED: {plan['title_skip_reason']}")
    elif plan["title"] is not None:
        out(f"  title (not pushed; use --push-title): {plan['title']!r}")


def print_result(result, out=print):
    out(f"--- applied offer_id {result['offer_id']} "
        f"(product_id {result['product_id']}) ---")
    for p in result.get("pictures", []):
        prim = " PRIMARY" if p.get("is_primary") else ""
        out(f"  picture {p.get('state')}{prim}: {p.get('url')}")
    if "title_task_id" in result:
        out(f"  title import task_id: {result['title_task_id']}  "
            f"status: {result.get('title_status')}")


# --------------------------------------------------------------------------- #
# Orchestration
# --------------------------------------------------------------------------- #
def run_push(
    client,
    rows,
    *,
    apply,
    limit,
    keep_existing,
    push_title,
    no_images=False,
    poll=1,
    out=print,
):
    """Plan (and optionally apply) the push for up to ``limit`` rows.

    ``client`` is any object exposing product_info_by_offer / pictures_import /
    product_attributes / product_import / product_import_info — the real
    OzonClient or a test fake. Returns (plans, results, confirmations).
    """
    touched = rows[: max(0, limit)] if limit is not None else rows
    offer_ids = [r.offer_id for r in touched]
    resolved = resolve_product_ids(client, offer_ids)

    mode = "APPLY (writing to Ozon)" if apply else "DRY-RUN (no API writes)"
    out("=== push-ozon ===")
    out(f"mode: {mode}   products: {len(touched)} (limit={limit})   "
        f"keep_existing={keep_existing}   push_title={push_title}")
    if no_images:
        out("note: --no-images set; image columns will NOT be pushed (title-only).")
    if push_title:
        out("WARNING: --push-title re-submits the whole product via "
            "/v3/product/import (changing only the name). Other attributes are "
            "preserved as fetched; products with thin attribute data are skipped.")
    out("")

    plans = []
    for row in touched:
        plan = plan_product(
            client, row, resolved,
            keep_existing=keep_existing, push_title=push_title,
        )
        plans.append(plan)
        print_plan(plan, out=out)
    out("")

    results = []
    confirmations = []
    if apply:
        for plan in plans:
            if not plan["resolved"]:
                continue
            result = apply_product(client, plan, poll=poll, no_images=no_images)
            results.append(result)
            print_result(result, out=out)
        out("")
        # Confirm the new primary image landed.
        for plan in plans:
            if not plan["resolved"] or not plan["primary"]:
                continue
            conf = confirm_product(client, plan)
            confirmations.append(conf)
            status = "OK" if conf["match"] else "MISMATCH"
            out(f"confirm {conf['offer_id']}: primary now "
                f"{conf['new_primary']!r} [{status}]")
        out("")

    # Summary
    n_resolved = sum(1 for p in plans if p["resolved"])
    n_imgs = sum(1 for p in plans if p["images_to_set"])
    n_titles = sum(1 for p in plans if p["title_push"])
    out(f"summary: {len(plans)} planned, {n_resolved} resolved, "
        f"{n_imgs} with images, {n_titles} title push(es); "
        f"{'WROTE to Ozon' if apply else 'NO writes (dry-run)'}.")
    return plans, results, confirmations


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #
def build_arg_parser():
    ap = argparse.ArgumentParser(
        prog="push-ozon",
        description=(
            "Push a relisted output .xlsx back to Ozon: set product images "
            "(safe, default) and optionally the title. Reads the xlsx only."
        ),
    )
    ap.add_argument("--in", dest="in_path", required=True,
                    help="the relisted output .xlsx to push from")
    ap.add_argument("--config", required=True,
                    help="fields.yaml mapping (e.g. fields.example.yaml)")
    ap.add_argument("--sheet", default=None,
                    help="sheet name (default: first processed/visible sheet)")
    ap.add_argument("--offer-col", default=None,
                    help="column LETTER or header KEYWORD for offer_id "
                         "(default: the config's key_column / Артикул / SKU)")
    ap.add_argument("--keep-existing", action="store_true",
                    help="append the product's current images AFTER the new ones")
    ap.add_argument("--push-title", action="store_true",
                    help="ALSO push the title (re-submits the product; off by default)")
    ap.add_argument("--no-images", action="store_true",
                    help="do NOT push images (e.g. title-only with --push-title)")
    grp = ap.add_mutually_exclusive_group()
    grp.add_argument("--dry-run", dest="apply", action="store_false", default=False,
                     help="(default) print the plan, make NO API writes")
    grp.add_argument("--apply", dest="apply", action="store_true",
                     help="perform the writes")
    ap.add_argument("--limit", type=int, default=1,
                    help="cap how many products are touched (default 1)")
    ap.add_argument("--poll", type=int, default=1,
                    help="title-import status polls after submit (default 1)")
    return ap


def main(argv=None) -> int:
    args = build_arg_parser().parse_args(argv)

    config = load_config(args.config)
    rows, layout = read_rows(
        args.in_path, config, sheet=args.sheet, offer_col_arg=args.offer_col
    )
    print(f"read {len(rows)} data row(s) from {os.path.abspath(args.in_path)}")
    print(f"  offer_col={layout['offer_col']}  image_cols={layout['image_cols']}  "
          f"title_col={layout['title_col']}  data_start_row={layout['data_start_row']}")
    print("")

    if not rows:
        print("no data rows found; nothing to push.")
        return 0

    # The client is only needed once we resolve/push. Build it lazily so a pure
    # --dry-run that still wants product_ids fails with a clear creds message.
    from ozon_excel_core.ozon_api import OzonClient

    client = OzonClient.from_env()

    run_push(
        client,
        rows,
        apply=args.apply,
        limit=args.limit,
        keep_existing=args.keep_existing,
        push_title=args.push_title,
        no_images=args.no_images,
        poll=args.poll,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
