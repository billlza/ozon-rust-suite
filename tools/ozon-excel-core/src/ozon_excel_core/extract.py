"""Read side: workbook -> list[ProductRow].

This is a content-only view used for inspection/diffing/logging. The writer does
NOT rebuild rows from these objects (see writer.py) — it mutates cells in place.
"""

from __future__ import annotations

from openpyxl import load_workbook

from . import images as images_mod
from .config import MappingConfig
from .header import detect_header
from .mapper import resolve_columns
from .model import ProductRow


def _processed_worksheets(wb, config: MappingConfig):
    out = []
    for ws in wb.worksheets:
        if getattr(ws, "sheet_state", "visible") != "visible":
            continue
        if config.matches_processed_sheet(ws.title):
            out.append(ws)
    return out


def extract(path, config: MappingConfig) -> dict:
    """Return {sheet_title: (MappedColumns, [ProductRow, ...])}."""
    path = str(path)
    keep_vba = path.lower().endswith(".xlsm")
    wb = load_workbook(path, data_only=False, keep_links=True, keep_vba=keep_vba)
    result: dict = {}
    try:
        for ws in _processed_worksheets(wb, config):
            header_block = detect_header(ws, config)
            mapped = resolve_columns(ws, config, header_block)
            rows = _extract_rows(ws, config, mapped)
            result[ws.title] = (mapped, rows)
    finally:
        wb.close()
    return result


def _extract_rows(ws, config: MappingConfig, mapped) -> list:
    rows: list = []
    start = mapped.header_block.data_start_row
    for r in range(start, (ws.max_row or start - 1) + 1):
        pr = ProductRow(sheet=ws.title, row=r)

        if mapped.key_column is not None:
            v = ws.cell(row=r, column=mapped.key_column.col_index).value
            pr.sku = None if v is None else str(v)

        if mapped.title is not None:
            v = ws.cell(row=r, column=mapped.title.col_index).value
            pr.title = None if v is None else str(v)
            pr.raw_targets[mapped.title.col_index] = v

        if mapped.listing is not None:
            v = ws.cell(row=r, column=mapped.listing.col_index).value
            pr.listing = None if v is None else str(v)
            pr.raw_targets[mapped.listing.col_index] = v

        for mc in mapped.images_main:
            cell = ws.cell(row=r, column=mc.col_index)
            spec = _field_for(config, "image_main")
            ic = images_mod.parse(cell, field_spec=spec, ws=ws)
            pr.images_main.append(ic)
            pr.raw_targets[mc.col_index] = cell.value

        for mc in mapped.images_additional:
            cell = ws.cell(row=r, column=mc.col_index)
            spec = _field_for(config, "image_additional")
            ic = images_mod.parse(cell, field_spec=spec, ws=ws)
            pr.images_additional.append(ic)
            pr.raw_targets[mc.col_index] = cell.value

        rows.append(pr)
    return rows


def _field_for(config: MappingConfig, role: str):
    if role == "image_main" and config.images_main:
        return config.images_main[0]
    if role == "image_additional" and config.images_additional:
        return config.images_additional[0]
    return None
