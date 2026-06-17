"""Multi-row header scan + data-start detection.

Implements the rule "do not assume row 1 = header / row 2 = data". The mapper
scans the top N rows, picks the row with the most keyword hits as the match
row, and finds where product data actually begins.
"""

from __future__ import annotations

import re

from .config import MappingConfig
from .errors import MappingError
from .model import HeaderBlock

_PRICE_LIKE_RE = re.compile(r"^\s*-?\d[\d\s.,]*\s*$")


def _normalized_row(ws, row_idx: int, normalize) -> list:
    """Return a list of (col_index, normalized_string) for non-empty cells."""
    out = []
    max_col = ws.max_column or 0
    for col in range(1, max_col + 1):
        v = ws.cell(row=row_idx, column=col).value
        if v is None:
            continue
        out.append((col, normalize(v), v))
    return out


def _count_keyword_hits(row_cells, config: MappingConfig) -> int:
    """How many configured keywords (any field) appear in this row's cells."""
    norm_strings = [s for (_c, s, _raw) in row_cells if s]
    if not norm_strings:
        return 0
    hits = 0
    for spec in config.all_field_specs():
        for kw in spec.keywords:
            for s in norm_strings:
                if _kw_in(kw.norm, s, spec.mode):
                    hits += 1
                    break
    return hits


def _kw_in(needle: str, haystack: str, mode: str) -> bool:
    if not needle:
        return False
    if mode == "equals":
        return needle == haystack
    if mode == "regex":
        try:
            return re.search(needle, haystack) is not None
        except re.error:
            return False
    return needle in haystack  # contains


def _string_cell_count(row_cells) -> int:
    return sum(1 for (_c, _s, raw) in row_cells if isinstance(raw, str) and raw.strip())


def _looks_like_data_value(raw) -> bool:
    """A value that looks like product data (not a header hint)."""
    if raw is None:
        return False
    if isinstance(raw, (int, float)):
        return True
    s = str(raw).strip()
    return s != ""


def detect_header(ws, config: MappingConfig) -> HeaderBlock:
    normalize = config.normalizer()
    scan_rows = max(1, config.header.scan_rows)
    last_scan = min(scan_rows, ws.max_row or scan_rows)

    rows = {r: _normalized_row(ws, r, normalize) for r in range(1, last_scan + 1)}

    # 1/2. Determine match_row.
    if config.header.match_row != "auto":
        match_row = int(config.header.match_row)
        if match_row < 1:
            raise MappingError(f"header.match_row {match_row} invalid")
    else:
        scored = []
        for r in range(1, last_scan + 1):
            hits = _count_keyword_hits(rows.get(r, []), config)
            scored.append((hits, _string_cell_count(rows.get(r, [])), -r, r))
        scored.sort(reverse=True)
        best = scored[0]
        if best[0] == 0:
            raise MappingError(
                f"header not found on sheet {ws.title!r}: no configured keyword "
                f"matched any of rows 1..{last_scan}. Pin header.match_row / "
                f"header.data_start in the config."
            )
        match_row = best[3]

    # 3. header_rows: contiguous block row 1..end_header where end_header is the
    #    greater of match_row and the last scanned "header-like" row.
    end_header = match_row
    for r in range(1, last_scan + 1):
        cells = rows.get(r, [])
        if not cells:
            continue
        n_str = _string_cell_count(cells)
        n_total = len(cells)
        n_numeric = sum(1 for (_c, _s, raw) in cells if isinstance(raw, (int, float)))
        n_pricey = sum(
            1
            for (_c, _s, raw) in cells
            if isinstance(raw, str) and _PRICE_LIKE_RE.match(raw)
        )
        header_like = (n_str >= max(1, n_total // 2)) and n_numeric == 0 and n_pricey == 0
        if header_like and r > end_header:
            end_header = r
    header_rows = list(range(1, end_header + 1))

    # 4. data_start_row.
    if config.header.data_start != "auto":
        data_start = int(config.header.data_start)
    else:
        data_start = _auto_data_start(ws, config, header_rows, normalize)

    # Guard.
    if data_start <= max(header_rows):
        from .errors import ConfigError

        raise ConfigError(
            f"data_start_row {data_start} must be > last header row "
            f"{max(header_rows)} on sheet {ws.title!r}"
        )

    return HeaderBlock(
        header_rows=header_rows,
        match_row=match_row,
        data_start_row=data_start,
    )


def _resolve_probe_col(ws, config: MappingConfig, match_row: int, normalize):
    """Pick the column whose non-empty value signals 'this is a data row':
    the key column if mapped, else the title column."""
    candidates = []
    if config.key_column is not None:
        candidates.append(config.key_column)
    if config.title is not None:
        candidates.append(config.title)

    header_cells = _normalized_row(ws, match_row, normalize)
    for spec in candidates:
        # explicit pins first
        if spec.index is not None:
            return spec.index
        if spec.letter is not None:
            from openpyxl.utils import column_index_from_string

            return column_index_from_string(spec.letter)
        for kw in spec.keywords:
            for (col, s, _raw) in header_cells:
                if _kw_in(kw.norm, s, spec.mode):
                    return col
    return None


def _auto_data_start(ws, config: MappingConfig, header_rows: list, normalize) -> int:
    start_floor = max(header_rows) + 1
    probe_col = _resolve_probe_col(ws, config, max(header_rows), normalize)
    max_row = ws.max_row or start_floor

    for r in range(start_floor, max_row + 1):
        if probe_col is not None:
            v = ws.cell(row=r, column=probe_col).value
            if _looks_like_data_value(v):
                return r
        else:
            # No probe col: first row with any non-empty cell below the header.
            for col in range(1, (ws.max_column or 1) + 1):
                if _looks_like_data_value(ws.cell(row=r, column=col).value):
                    return r
    # No data rows found; data conceptually starts right after header.
    return start_floor
