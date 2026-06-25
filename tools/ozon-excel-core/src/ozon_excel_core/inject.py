"""Inject per-row product data into the mapped target cells of a template.

This promotes scripts/demo_real.py:_inject_products into the package so the
sidecar (Module 4 export/delivery) can call it through the CLI. It writes each
row's {title, listing, primary_image, additional_images[]} into EXACTLY the
title / listing / images_main[0] / images_additional cells the mapper resolves,
mirroring demo_real.py's cell logic precisely:

  - title      -> the mapped title column cell
  - listing    -> the mapped listing column cell
  - primary    -> images_main[0]
  - additional -> distributed across images_additional columns: the first
                  (multi_url) column takes everything left, newline-joined; any
                  trailing single-url columns take one image each.

Images are PUBLIC URL strings — never embedded rasters.
"""

from __future__ import annotations

from openpyxl import load_workbook

from .config import MappingConfig
from .header import detect_header
from .mapper import resolve_columns

# The single product sheet the template + fields.example.yaml describe.
MAIN_SHEET = "Шаблон для поставщика"


def _first_url(v):
    """Ozon image fields can be a str or a list of urls; cell values must be
    scalar strings."""
    if isinstance(v, (list, tuple)):
        v = v[0] if v else None
    return str(v) if v else None


def _letter(mc):
    return mc.column_letter if mc is not None else None


def inject_rows(in_xlsx, out_xlsx, config: MappingConfig, rows, *, sheet=None):
    """Overwrite the first ``len(rows)`` data rows of the template with ``rows``.

    Each row is a mapping with keys:
      - title              (str)
      - listing            (str)
      - primary_image      (str | None)
      - additional_images  (list[str])

    Returns a dict describing where each piece landed (for the report). Mirrors
    demo_real.py:_inject_products cell logic precisely.
    """
    wb = load_workbook(str(in_xlsx))
    ws = wb[sheet or MAIN_SHEET]
    hb = detect_header(ws, config)
    mapped = resolve_columns(ws, config, hb)

    title_col = mapped.title.col_index if mapped.title else None
    listing_col = mapped.listing.col_index if mapped.listing else None
    main_col = mapped.images_main[0].col_index if mapped.images_main else None
    add_cols = [mc.col_index for mc in mapped.images_additional]

    data_start = hb.data_start_row
    placed = {
        "data_start_row": data_start,
        "title_col": _letter(mapped.title),
        "listing_col": _letter(mapped.listing),
        "main_image_col": _letter(mapped.images_main[0]) if mapped.images_main else None,
        "additional_image_cols": [_letter(mc) for mc in mapped.images_additional],
        "rows": [],
    }

    for i, row in enumerate(rows):
        r = data_start + i
        title = str(row.get("title") or "")
        listing = str(row.get("listing") or "")
        if title_col:
            ws.cell(row=r, column=title_col, value=title)
        if listing_col:
            ws.cell(row=r, column=listing_col, value=listing)

        primary = _first_url(row.get("primary_image"))
        if main_col and primary:
            ws.cell(row=r, column=main_col, value=primary)

        # Distribute the remaining images across the additional target columns.
        # The first additional column is multi_url (accepts several, newline-
        # joined); any further single-url columns (e.g. "photo 2") take one each.
        others = [
            str(u)
            for u in (row.get("additional_images") or [])
            if u and str(u) != primary
        ]
        if add_cols and others:
            first, rest = add_cols[0], add_cols[1:]
            # The trailing single-url columns each take one image; the multi_url
            # first column takes everything that's left, newline-joined.
            n_singles = min(len(rest), len(others))
            singles = others[len(others) - n_singles:] if n_singles else []
            multi = others[: len(others) - n_singles]
            if multi:
                ws.cell(row=r, column=first, value="\n".join(multi))
            elif singles:
                # No leftover for the multi column; still fill it with one image
                # so the main multi_url slot is not left empty.
                ws.cell(row=r, column=first, value=singles[0])
                singles = singles[1:]
                rest = rest[1:]
            for col, url in zip(rest, singles):
                ws.cell(row=r, column=col, value=url)

        placed["rows"].append({"row": r, "title": title[:48]})

    # Drop any leftover sample rows so only the supplied rows are processed
    # (un-injected sample rows still carry placeholder urls/values).
    last = data_start + len(rows) - 1
    if rows and ws.max_row > last:
        ws.delete_rows(last + 1, ws.max_row - last)

    wb.save(str(out_xlsx))
    wb.close()
    return placed
