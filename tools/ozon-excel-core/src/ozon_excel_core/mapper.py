"""Resolve header keywords / explicit letters / indices into target columns.

Resolution precedence (per field):
    1. explicit index (1-based)
    2. explicit letter
    3. keywords against the header match_row

Single-valued fields (title/listing) that match multiple columns -> MappingError.
Image fields may legitimately resolve to multiple columns.
Everything not resolved is FROZEN.
"""

from __future__ import annotations

import re
from typing import Optional

from openpyxl.utils import column_index_from_string, get_column_letter

from .config import FieldSpec, MappingConfig
from .errors import MappingError
from .model import (
    ColumnRole,
    HeaderBlock,
    MappedColumn,
    MappedColumns,
)

_ROLE_ENUM = {
    "title": ColumnRole.TITLE,
    "listing": ColumnRole.LISTING,
    "image_main": ColumnRole.IMAGE_MAIN,
    "image_additional": ColumnRole.IMAGE_ADDITIONAL,
    "key": ColumnRole.FROZEN,  # key column is frozen (label only)
}


def _kw_match(needle: str, haystack: str, mode: str) -> bool:
    if not needle:
        return False
    if mode == "equals":
        return needle == haystack
    if mode == "regex":
        try:
            return re.search(needle, haystack) is not None
        except re.error:
            return False
    return needle in haystack


def _header_cells(ws, match_row: int, normalize) -> list:
    """List of (col_index, normalized_string) for non-empty header cells."""
    out = []
    for col in range(1, (ws.max_column or 0) + 1):
        v = ws.cell(row=match_row, column=col).value
        if v is None:
            continue
        out.append((col, normalize(v)))
    return out


def _resolve_one(spec: FieldSpec, header_cells, mode_default: str):
    """Return a list of (col_index, matched_by, matched_token) for a field.
    Explicit index/letter short-circuit; otherwise collect all keyword hits."""
    if spec.index is not None:
        return [(spec.index, "index", None)]
    if spec.letter is not None:
        return [(column_index_from_string(spec.letter), "letter", None)]

    hits = []
    seen_cols = set()
    for kw in spec.keywords:
        for (col, s) in header_cells:
            if col in seen_cols:
                continue
            if _kw_match(kw.norm, s, spec.mode):
                hits.append((col, "keyword", kw.raw))
                seen_cols.add(col)
    return hits


def _to_mapped(spec: FieldSpec, sheet: str, col_index: int, matched_by: str,
               matched_token, role: ColumnRole) -> MappedColumn:
    return MappedColumn(
        role=role,
        sheet=sheet,
        col_index=col_index,
        column_letter=get_column_letter(col_index),
        matched_by=matched_by,
        matched_token=matched_token,
        image_form_hint=spec.image_form,
    )


def resolve_columns(ws, config: MappingConfig, header_block: HeaderBlock) -> MappedColumns:
    normalize = config.normalizer()
    header_cells = _header_cells(ws, header_block.match_row, normalize)
    sheet = ws.title

    # Track which letter pins to which role to catch contradictions.
    pin_letters: dict = {}

    def note_pin(spec: FieldSpec, role_name: str):
        if spec.letter is not None:
            prev = pin_letters.get(spec.letter)
            if prev is not None and prev != role_name:
                raise MappingError(
                    f"column letter {spec.letter!r} pinned to both {prev!r} and "
                    f"{role_name!r}"
                )
            pin_letters[spec.letter] = role_name

    # --- title (single-valued) ---
    title_mc = None
    if config.title is not None:
        note_pin(config.title, "title")
        hits = _resolve_one(config.title, header_cells, config.match.mode)
        title_mc = _single(
            config.title, sheet, hits, ColumnRole.TITLE, "title", config
        )

    # --- listing (single-valued) ---
    listing_mc = None
    if config.listing is not None:
        note_pin(config.listing, "listing")
        hits = _resolve_one(config.listing, header_cells, config.match.mode)
        listing_mc = _single(
            config.listing, sheet, hits, ColumnRole.LISTING, "listing", config
        )

    # --- images_main (multi-valued across specs, single-valued per spec) ---
    images_main = []
    for spec in config.images_main:
        note_pin(spec, "images_main")
        hits = _resolve_one(spec, header_cells, config.match.mode)
        col, by, tok = _single_hit(spec, sheet, hits, "images_main")
        if col is not None:
            images_main.append(
                _to_mapped(spec, sheet, col, by, tok, ColumnRole.IMAGE_MAIN)
            )

    # --- images_additional (multi-valued across specs, single-valued per spec) ---
    images_additional = []
    for spec in config.images_additional:
        note_pin(spec, "images_additional")
        hits = _resolve_one(spec, header_cells, config.match.mode)
        col, by, tok = _single_hit(spec, sheet, hits, "images_additional")
        if col is not None:
            images_additional.append(
                _to_mapped(spec, sheet, col, by, tok, ColumnRole.IMAGE_ADDITIONAL)
            )

    # --- key column (label only; frozen) ---
    key_mc = None
    if config.key_column is not None:
        hits = _resolve_one(config.key_column, header_cells, config.match.mode)
        if hits:
            col, by, tok = hits[0]
            key_mc = _to_mapped(config.key_column, sheet, col, by, tok, ColumnRole.FROZEN)

    # --- required check ---
    _check_required(config, title_mc, listing_mc, images_main, images_additional)

    # --- target set (key column NOT included; it is frozen) ---
    target_indices = set()
    for mc in ([title_mc, listing_mc] + images_main + images_additional):
        if mc is not None:
            target_indices.add(mc.col_index)

    # Detect a target column accidentally landing on the key column.
    if key_mc is not None and key_mc.col_index in target_indices:
        raise MappingError(
            f"key column {key_mc.column_letter} on sheet {sheet!r} also resolved "
            f"as an editable target; refusing to write the SKU column."
        )

    return MappedColumns(
        sheet=sheet,
        header_block=header_block,
        title=title_mc,
        listing=listing_mc,
        images_main=images_main,
        images_additional=images_additional,
        target_col_indices=frozenset(target_indices),
        key_column=key_mc,
    )


def _single_hit(spec: FieldSpec, sheet, hits, name: str):
    """Image-field ambiguity guard. Each individual image FieldSpec may resolve
    to at most ONE physical column. Multiple image columns are expressed by
    supplying multiple specs (one per column), each matching a single column —
    NOT by a single keyword spec over-matching several columns under
    mode:contains.

    Returns ``(col, matched_by, matched_token)`` for the single hit, or
    ``(None, None, None)`` when the spec resolved nothing. Raises MappingError —
    fail-closed — when a keyword spec matches >1 column, exactly like the
    single-valued title/listing fields. This is what stops a frozen physical
    column whose header substring-contains an image keyword (e.g.
    'Главное фото для внутреннего архива' ⊃ 'Главное фото') from being silently
    added to the write set."""
    if not hits:
        return (None, None, None)
    if len(hits) > 1:
        cols = ", ".join(f"{get_column_letter(c)}({tok!r})" for (c, _b, tok) in hits)
        raise MappingError(
            f"image field {name!r} on sheet {sheet!r} matched multiple columns: "
            f"{cols}. A keyword image column must resolve to exactly one column; "
            f"pin each with letter/index, tighten the keyword, or use mode:equals."
        )
    return hits[0]


def _single(spec: FieldSpec, sheet, hits, role: ColumnRole, name: str, config) -> Optional[MappedColumn]:
    if not hits:
        return None
    if len(hits) > 1:
        cols = ", ".join(f"{get_column_letter(c)}({tok!r})" for (c, _b, tok) in hits)
        raise MappingError(
            f"field {name!r} on sheet {sheet!r} matched multiple columns: {cols}. "
            f"Pin it with letter/index in the config."
        )
    col, by, tok = hits[0]
    return _to_mapped(spec, sheet, col, by, tok, role)


def _check_required(config, title_mc, listing_mc, images_main, images_additional):
    missing = []
    if config.title is not None and config.title.required and title_mc is None:
        missing.append("title")
    if config.listing is not None and config.listing.required and listing_mc is None:
        missing.append("listing")
    for spec in config.images_main:
        if spec.required and not images_main:
            missing.append("images_main")
            break
    for spec in config.images_additional:
        if spec.required and not images_additional:
            missing.append("images_additional")
            break

    if missing:
        if config.policy.on_missing_required == "error":
            raise MappingError(
                f"required column(s) unresolved: {', '.join(missing)}. "
                f"Adjust keywords or pin letter/index."
            )
        # warn policy: caller may inspect; we just allow None.
