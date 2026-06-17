"""Sample generator: build a representative Ozon-style workbook.

Faithful structural stand-in (multi-row header, ~50 columns, multiple image
forms, hidden + instruction sheets, styled/merged/validated frozen cells). Never
claims byte-identity with a real Ozon file. Uses only openpyxl create-APIs
(allowed for creation).
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

MAIN_SHEET = "Шаблон для поставщика"
VALIDATION_SHEET = "validation"
INSTRUCTION_SHEET = "Инструкция"

# Column plan: (row1 human label, row2 technical key, row3 hint, role)
# roles: frozen | title | listing | img_hyperlink | img_multi | img_plain | key
def _column_plan():
    cols = []
    # key + content + images first, then physical/logistics, then filler.
    cols.append(("Артикул*", "offer_id", "Уникальный код товара", "key"))
    cols.append(("Название товара", "name", "До 200 символов", "title"))
    cols.append(("Описание", "annotation", "Текст описания товара", "listing"))
    cols.append(("Ссылка на главное фото", "main_photo", "Ссылка =HYPERLINK", "img_hyperlink"))
    cols.append(("Ссылки на дополнительные фото", "additional_photos", "Несколько ссылок", "img_multi"))
    cols.append(("Ссылка на фото 2", "photo_2", "Одна ссылка", "img_plain"))

    # Physical / logistics (frozen)
    physical = [
        ("Цена, руб", "price", "Число", "frozen_price"),
        ("Цена до скидки", "old_price", "Число", "frozen_price"),
        ("НДС, %", "vat", "0/10/20", "frozen"),
        ("Штрихкод", "barcode", "EAN-13", "frozen"),
        ("Остаток", "stock", "Целое", "frozen_int"),
        ("Длина, мм", "depth", "Целое", "frozen_int"),
        ("Ширина, мм", "width", "Целое", "frozen_int"),
        ("Высота, мм", "height", "Целое", "frozen_int"),
        ("Вес, г", "weight", "Целое", "frozen_int"),
        ("Материал", "material", "Текст", "frozen"),
        ("Бренд", "brand", "Текст", "frozen"),
        ("Партномер", "part_number", "Текст", "frozen"),
        ("Тип", "type", "Текст", "frozen"),
        ("Способ доставки", "delivery", "FBO/FBS", "frozen_dv"),
    ]
    cols.extend(physical)

    # Filler category-characteristic columns to reach ~50 total.
    n_filler = 50 - len(cols)
    for i in range(1, n_filler + 1):
        cols.append(
            (f"Характеристика {i}", f"attr_{i}", "Текст", "frozen")
        )
    return cols


PRODUCTS = [
    {
        "offer_id": "SKU-1001",
        "name": "  чайник  электрический  стеклянный  ",
        "annotation": "Объём 1.7 л.\r\nМощность 2200 Вт.\n\n\nГарантия 12 месяцев.",
        "main_photo": '=HYPERLINK("https://img.ozon.test/1001/main.jpg","фото")',
        "additional_photos": "https://img.ozon.test/1001/a1.jpg\nhttps://img.ozon.test/1001/a2.jpg\nhttps://img.ozon.test/1001/a3.jpg",
        "photo_2": "https://img.ozon.test/1001/p2.jpg",
        "price": 2490.0,
        "old_price": 3200.0,
        "vat": 20,
        "barcode": "4600123456789",
        "stock": 37,
        "depth": 250,
        "width": 180,
        "height": 230,
        "weight": 1100,
        "material": "Стекло",
        "brand": "ТеплоДом",
        "part_number": "TD-1700",
        "type": "Чайник",
        "delivery": "FBO",
    },
    {
        "offer_id": "SKU-1002",
        "name": "Кружка керамическая 350 мл, синяя",
        "annotation": "Подходит для микроволновой печи.",
        "main_photo": '=HYPERLINK("https://img.ozon.test/1002/main.jpg","главное фото")',
        "additional_photos": "https://img.ozon.test/1002/a1.jpg,https://img.ozon.test/1002/a2.jpg",
        "photo_2": "https://img.ozon.test/1002/p2.jpg",
        "price": 390.0,
        "old_price": 520.0,
        "vat": 20,
        "barcode": "4600123456796",
        "stock": 120,
        "depth": 120,
        "width": 90,
        "height": 95,
        "weight": 300,
        "material": "Керамика",
        "brand": "ПосудаПлюс",
        "part_number": "PP-350B",
        "type": "Кружка",
        "delivery": "FBS",
    },
    {
        "offer_id": "SKU-1003",
        "name": "Нож кухонный шеф 20 см",
        "annotation": "Нержавеющая сталь. Эргономичная ручка.",
        "main_photo": '=HYPERLINK("https://img.ozon.test/1003/main.jpg")',
        "additional_photos": "https://img.ozon.test/1003/a1.jpg https://img.ozon.test/1003/a2.jpg",
        "photo_2": "https://img.ozon.test/1003/p2.jpg",
        "price": 1290.0,
        "old_price": 1290.0,
        "vat": 20,
        "barcode": "4600123456802",
        "stock": 54,
        "depth": 330,
        "width": 50,
        "height": 25,
        "weight": 220,
        "material": "Сталь",
        "brand": "ОстрыйКрай",
        "part_number": "OK-CHEF20",
        "type": "Нож",
        "delivery": "FBO",
    },
    {
        "offer_id": "SKU-1004",
        "name": "Полотенце махровое 70x140, белое",
        "annotation": "Хлопок 100%. Плотность 450 г/м².",
        "main_photo": '=HYPERLINK("https://img.ozon.test/1004/main.jpg","фото товара")',
        "additional_photos": "https://img.ozon.test/1004/a1.jpg",
        "photo_2": "https://img.ozon.test/1004/p2.jpg",
        "price": 690.0,
        "old_price": 890.0,
        "vat": 10,
        "barcode": "4600123456819",
        "stock": 200,
        "depth": 250,
        "width": 200,
        "height": 40,
        "weight": 450,
        "material": "Хлопок",
        "brand": "МягкийДом",
        "part_number": "MD-70140W",
        "type": "Полотенце",
        "delivery": "FBS",
    },
    {
        "offer_id": "SKU-1005",
        "name": "Лампа настольная LED, чёрная",
        "annotation": "3 режима яркости. USB-питание.",
        "main_photo": '=HYPERLINK("https://img.ozon.test/1005/main.jpg","фото")',
        "additional_photos": "https://img.ozon.test/1005/a1.jpg\nhttps://img.ozon.test/1005/a2.jpg",
        "photo_2": "https://img.ozon.test/1005/p2.jpg",
        "price": 1590.0,
        "old_price": 1990.0,
        "vat": 20,
        "barcode": "4600123456826",
        "stock": 75,
        "depth": 150,
        "width": 120,
        "height": 400,
        "weight": 600,
        "material": "Пластик",
        "brand": "СветПро",
        "part_number": "SP-LED01",
        "type": "Лампа",
        "delivery": "FBO",
    },
]


def gen_sample(out_path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = MAIN_SHEET

    plan = _column_plan()
    n_cols = len(plan)
    data_start = 4

    header_fill = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # --- header rows (1..3) ---
    for ci, (label, key, hint, role) in enumerate(plan, start=1):
        c1 = ws.cell(row=1, column=ci, value=label)
        c2 = ws.cell(row=2, column=ci, value=key)
        c3 = ws.cell(row=3, column=ci, value=hint)
        for c in (c1, c2, c3):
            c.fill = header_fill
            c.font = header_font
            c.alignment = center

    # --- representative formatting on frozen cells ---
    role_by_col = {ci: role for ci, (_l, _k, _h, role) in enumerate(plan, start=1)}
    col_index = {key: ci for ci, (_l, key, _h, _r) in enumerate(plan, start=1)}

    # one column width override (on Описание)
    ws.column_dimensions[get_column_letter(col_index["annotation"])].width = 48
    # a couple merged header cells: merge a filler attr label across rows 1..1? Use
    # a horizontal merge of two filler labels on row 1 to exercise merge preservation.
    last = n_cols
    merge_a = get_column_letter(last - 1)
    merge_b = get_column_letter(last)
    ws.merge_cells(f"{merge_a}1:{merge_b}1")
    ws.cell(row=1, column=last - 1, value="Прочие характеристики")

    # --- product data rows (start at row 4) ---
    for ri, product in enumerate(PRODUCTS, start=data_start):
        for ci, (_label, key, _hint, role) in enumerate(plan, start=1):
            val = product.get(key)
            if val is None:
                continue
            cell = ws.cell(row=ri, column=ci, value=val)
            if role == "frozen_price":
                cell.number_format = "#,##0.00"
            elif role == "frozen_int":
                cell.number_format = "0"

    # --- validation sheet (hidden) ---
    vws = wb.create_sheet(VALIDATION_SHEET)
    delivery_opts = ["FBO", "FBS"]
    vat_opts = [0, 10, 20]
    vws["A1"] = "delivery"
    for i, opt in enumerate(delivery_opts, start=2):
        vws.cell(row=i, column=1, value=opt)
    vws["B1"] = "vat"
    for i, opt in enumerate(vat_opts, start=2):
        vws.cell(row=i, column=2, value=opt)
    vws.sheet_state = "hidden"

    # data validation on the frozen delivery column, backed by the validation sheet
    delivery_letter = get_column_letter(col_index["delivery"])
    dv = DataValidation(
        type="list",
        formula1=f"={VALIDATION_SHEET}!$A$2:$A$3",
        allow_blank=True,
    )
    dv.add(f"{delivery_letter}{data_start}:{delivery_letter}{data_start + len(PRODUCTS) - 1}")
    ws.add_data_validation(dv)

    # --- instruction sheet ---
    iws = wb.create_sheet(INSTRUCTION_SHEET)
    iws["A1"] = "Инструкция по заполнению шаблона"
    iws["A1"].font = Font(bold=True, size=14)
    iws["A3"] = "1. Заполните название и описание товара."
    iws["A4"] = "2. Укажите ссылки на фотографии."
    iws["A5"] = "3. Не изменяйте служебные строки (1-3)."

    wb.save(str(out_path))
    wb.close()
