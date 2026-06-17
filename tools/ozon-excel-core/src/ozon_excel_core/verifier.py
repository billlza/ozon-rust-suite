"""Verifier: prove only mapped target columns changed.

Independent of the writer: it re-derives the allowed set from the same config so
it is a real check, not a replay. Walks every cell of every sheet, compares a
signature (value, number_format, hyperlink_key), and flags any frozen/header/
structural change as unexpected. CLI exits non-zero when ok is False.
"""

from __future__ import annotations

from dataclasses import dataclass, field

from openpyxl import load_workbook

from .config import MappingConfig
from .errors import MappingError
from .header import detect_header
from .mapper import resolve_columns


@dataclass(frozen=True)
class CellDiff:
    sheet: str
    coordinate: str
    field: str  # value | number_format | hyperlink
    old: object
    new: object
    role: str  # title | listing | image | frozen | header | structural


@dataclass
class VerifyReport:
    in_path: str
    out_path: str
    ok: bool = True
    expected_changes: list = field(default_factory=list)  # list[CellDiff]
    unexpected_changes: list = field(default_factory=list)  # list[CellDiff]
    summary: dict = field(default_factory=dict)
    notes: list = field(default_factory=list)

    def to_text(self) -> str:
        lines = []
        lines.append("=== ozon-excel-core verify report ===")
        lines.append(f"in  : {self.in_path}")
        lines.append(f"out : {self.out_path}")
        lines.append(
            "NOTE: openpyxl cannot round-trip embedded images/charts; their presence "
            "is caught by preflight, not here."
        )
        lines.append("")
        lines.append(f"RESULT: {'OK — only mapped content columns changed' if self.ok else 'FAIL — unexpected changes detected'}")
        lines.append("")
        lines.append("-- summary --")
        for k, v in sorted(self.summary.items()):
            lines.append(f"  {k}: {v}")
        lines.append("")
        lines.append(f"-- expected changes ({len(self.expected_changes)}) --")
        for d in self.expected_changes:
            lines.append(
                f"  {d.sheet}!{d.coordinate} [{d.role}/{d.field}]: "
                f"{_short(d.old)} -> {_short(d.new)}"
            )
        if self.unexpected_changes:
            lines.append("")
            lines.append(f"-- UNEXPECTED changes ({len(self.unexpected_changes)}) --")
            for d in self.unexpected_changes:
                lines.append(
                    f"  {d.sheet}!{d.coordinate} [{d.role}/{d.field}]: "
                    f"{_short(d.old)} -> {_short(d.new)}"
                )
        return "\n".join(lines)

    def to_json_dict(self) -> dict:
        return {
            "in_path": self.in_path,
            "out_path": self.out_path,
            "ok": self.ok,
            "summary": self.summary,
            "expected_changes": [_diff_dict(d) for d in self.expected_changes],
            "unexpected_changes": [_diff_dict(d) for d in self.unexpected_changes],
            "notes": self.notes,
        }


def _diff_dict(d: CellDiff) -> dict:
    return {
        "sheet": d.sheet,
        "coordinate": d.coordinate,
        "field": d.field,
        "role": d.role,
        "old": _jsonable(d.old),
        "new": _jsonable(d.new),
    }


def _jsonable(v):
    if v is None or isinstance(v, (str, int, float, bool)):
        return v
    return str(v)


def _short(v) -> str:
    s = "<empty>" if v is None or v == "" else repr(v)
    if len(s) > 80:
        s = s[:77] + "..."
    return s


def _hyperlink_key(cell):
    hl = cell.hyperlink
    if hl is None:
        return None
    return (getattr(hl, "target", None), getattr(hl, "location", None))


def _cell_signature(cell):
    return (cell.value, cell.number_format, _hyperlink_key(cell))


def _used_cell_coords(ws):
    """All coordinates worth comparing on a sheet: union of cells that have a
    value, a style, a hyperlink, or a non-default number format."""
    coords = set()
    for row in ws.iter_rows():
        for cell in row:
            if (
                cell.value is not None
                or cell.hyperlink is not None
                or (cell.has_style and cell.number_format not in (None, "General"))
                or cell.has_style
            ):
                coords.add((cell.row, cell.column))
    return coords


def _build_allowed(path, config: MappingConfig) -> dict:
    """Return {sheet_title: (allowed_set, role_by_col, data_start)} where
    allowed_set is {(row, col)} for editable target cells in data rows."""
    keep_vba = str(path).lower().endswith(".xlsm")
    wb = load_workbook(str(path), data_only=False, keep_links=True, keep_vba=keep_vba)
    out = {}
    try:
        for ws in wb.worksheets:
            if getattr(ws, "sheet_state", "visible") != "visible":
                out[ws.title] = (set(), {}, None, set())
                continue
            if not config.matches_processed_sheet(ws.title):
                out[ws.title] = (set(), {}, None, set())
                continue
            hb = detect_header(ws, config)
            mapped = resolve_columns(ws, config, hb)
            role_by_col = {}
            if mapped.title:
                role_by_col[mapped.title.col_index] = "title"
            if mapped.listing:
                role_by_col[mapped.listing.col_index] = "listing"
            for mc in mapped.images_main:
                role_by_col[mc.col_index] = "image"
            for mc in mapped.images_additional:
                role_by_col[mc.col_index] = "image"
            target_cols = set(role_by_col)
            header_rows = set(hb.header_rows)
            out[ws.title] = (target_cols, role_by_col, hb.data_start_row, header_rows)
    finally:
        wb.close()
    return out


def verify(in_path, out_path, config: MappingConfig) -> VerifyReport:
    in_path = str(in_path)
    out_path = str(out_path)
    report = VerifyReport(in_path=in_path, out_path=out_path)

    # Derive the allowed (editable) target set from the INPUT mapping. If the
    # mapping is ambiguous (e.g. an image keyword over-matches a frozen physical
    # column), the mapper refuses — and so must the verifier: an unresolvable
    # mapping is reported as a definitive FAIL, never silently treated as OK.
    try:
        allowed = _build_allowed(in_path, config)
    except MappingError as exc:
        report.ok = False
        report.unexpected_changes.append(
            CellDiff("<config>", "-", "mapping", "resolvable", str(exc), "structural")
        )
        report.summary = {
            "expected_changes": 0,
            "unexpected_changes": len(report.unexpected_changes),
            "frozen_cells_compared": 0,
            "sheets_compared": 0,
        }
        report.notes.append(
            "Mapping could not be resolved unambiguously; cannot prove the "
            "frozen set. Treated as FAIL (fail-closed)."
        )
        return report

    wb_a = load_workbook(in_path, data_only=False, keep_links=True)
    wb_b = load_workbook(out_path, data_only=False, keep_links=True)

    try:
        names_a = set(wb_a.sheetnames)
        names_b = set(wb_b.sheetnames)
        if names_a != names_b:
            report.ok = False
            for missing in sorted(names_a - names_b):
                report.unexpected_changes.append(
                    CellDiff(missing, "-", "sheet", "present", "removed", "structural")
                )
            for added in sorted(names_b - names_a):
                report.unexpected_changes.append(
                    CellDiff(added, "-", "sheet", "absent", "added", "structural")
                )

        n_frozen_compared = 0
        n_expected = 0

        for sheet in sorted(names_a & names_b):
            ws_a = wb_a[sheet]
            ws_b = wb_b[sheet]
            target_cols, role_by_col, data_start, header_rows = allowed.get(
                sheet, (set(), {}, None, set())
            )

            coords = _used_cell_coords(ws_a) | _used_cell_coords(ws_b)
            for (row, col) in sorted(coords):
                ca = ws_a.cell(row=row, column=col)
                cb = ws_b.cell(row=row, column=col)
                sig_a = _cell_signature(ca)
                sig_b = _cell_signature(cb)
                if sig_a == sig_b:
                    # Still counts toward the frozen-compared tally if not a target.
                    if not _is_target_cell(col, row, target_cols, data_start, header_rows):
                        n_frozen_compared += 1
                    continue

                is_target = _is_target_cell(col, row, target_cols, data_start, header_rows)
                # produce a diff per differing field
                for field_name, old_v, new_v in _diff_fields(sig_a, sig_b):
                    if is_target:
                        role = role_by_col.get(col, "title")
                        report.expected_changes.append(
                            CellDiff(sheet, ca.coordinate, field_name, old_v, new_v, role)
                        )
                        n_expected += 1
                    else:
                        role = "header" if row in header_rows else "frozen"
                        report.unexpected_changes.append(
                            CellDiff(sheet, ca.coordinate, field_name, old_v, new_v, role)
                        )
                if not is_target:
                    report.ok = False
                else:
                    pass
                if not is_target:
                    n_frozen_compared += 1

            # structural checks per sheet
            _check_structural(sheet, ws_a, ws_b, report)

        # defined names (workbook-level)
        _check_defined_names(wb_a, wb_b, report)

        report.summary = {
            "expected_changes": n_expected,
            "unexpected_changes": len(report.unexpected_changes),
            "frozen_cells_compared": n_frozen_compared,
            "sheets_compared": len(names_a & names_b),
        }
        if report.unexpected_changes:
            report.ok = False
    finally:
        wb_a.close()
        wb_b.close()

    return report


def _is_target_cell(col, row, target_cols, data_start, header_rows) -> bool:
    if col not in target_cols:
        return False
    if data_start is None:
        return False
    if row < data_start:
        return False
    if row in header_rows:
        return False
    return True


def _diff_fields(sig_a, sig_b):
    out = []
    va, fa, ha = sig_a
    vb, fb, hb = sig_b
    if va != vb:
        out.append(("value", va, vb))
    if fa != fb:
        out.append(("number_format", fa, fb))
    if ha != hb:
        out.append(("hyperlink", ha, hb))
    return out


def _check_structural(sheet, ws_a, ws_b, report: VerifyReport) -> None:
    # merged cells
    merges_a = {str(r) for r in ws_a.merged_cells.ranges}
    merges_b = {str(r) for r in ws_b.merged_cells.ranges}
    if merges_a != merges_b:
        report.ok = False
        report.unexpected_changes.append(
            CellDiff(sheet, "-", "merged_cells", sorted(merges_a), sorted(merges_b), "structural")
        )

    # column widths
    widths_a = {k: v.width for k, v in ws_a.column_dimensions.items() if v.width is not None}
    widths_b = {k: v.width for k, v in ws_b.column_dimensions.items() if v.width is not None}
    if widths_a != widths_b:
        report.ok = False
        report.unexpected_changes.append(
            CellDiff(sheet, "-", "column_widths", widths_a, widths_b, "structural")
        )

    # row heights
    heights_a = {k: v.height for k, v in ws_a.row_dimensions.items() if v.height is not None}
    heights_b = {k: v.height for k, v in ws_b.row_dimensions.items() if v.height is not None}
    if heights_a != heights_b:
        report.ok = False
        report.unexpected_changes.append(
            CellDiff(sheet, "-", "row_heights", heights_a, heights_b, "structural")
        )

    # data validations (sqref sets)
    dv_a = {str(dv.sqref) for dv in ws_a.data_validations.dataValidation}
    dv_b = {str(dv.sqref) for dv in ws_b.data_validations.dataValidation}
    if dv_a != dv_b:
        report.ok = False
        report.unexpected_changes.append(
            CellDiff(sheet, "-", "data_validations", sorted(dv_a), sorted(dv_b), "structural")
        )


def _check_defined_names(wb_a, wb_b, report: VerifyReport) -> None:
    try:
        names_a = {n: dn.value for n, dn in wb_a.defined_names.items()}
        names_b = {n: dn.value for n, dn in wb_b.defined_names.items()}
    except AttributeError:
        # older openpyxl returned a list-like; normalize
        names_a = {dn.name: dn.value for dn in wb_a.defined_names.definedName}
        names_b = {dn.name: dn.value for dn in wb_b.defined_names.definedName}
    if names_a != names_b:
        report.ok = False
        report.unexpected_changes.append(
            CellDiff("<workbook>", "-", "defined_names", names_a, names_b, "structural")
        )
