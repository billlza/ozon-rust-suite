"""Surgical writer: load once, mutate ONLY mapped target cells, save a new file.

Never mutates the input. Never rebuilds sheets, never copies cell-by-cell, never
writes the input path. Frozen columns and non-processed sheets are never
addressed.
"""

from __future__ import annotations

import os
from dataclasses import dataclass, field

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from . import images as images_mod
from .config import MappingConfig
from .errors import ConfigError, PreflightError
from .header import detect_header
from .mapper import resolve_columns
from .model import CellRef, ColumnRole, ImageForm
from .preflight import preflight


@dataclass
class ProcessResult:
    in_path: str
    out_path: str
    rows_seen: int = 0
    changed_by_role: dict = field(default_factory=dict)  # role name -> count
    changed_cells: list = field(default_factory=list)  # list[CellRef]
    preflight_warnings: list = field(default_factory=list)  # list[Risk]
    mapping_summary: list = field(default_factory=list)  # list[str]
    skipped_embedded: list = field(default_factory=list)  # list[CellRef]
    skipped_merged: list = field(default_factory=list)  # list[CellRef]

    def total_changed(self) -> int:
        return len(self.changed_cells)


def _processed_worksheets(wb, config: MappingConfig):
    out = []
    for ws in wb.worksheets:
        if getattr(ws, "sheet_state", "visible") != "visible":
            continue
        if config.matches_processed_sheet(ws.title):
            out.append(ws)
    return out


def _field_for_role(config: MappingConfig, role: ColumnRole):
    if role is ColumnRole.IMAGE_MAIN and config.images_main:
        return config.images_main[0]
    if role is ColumnRole.IMAGE_ADDITIONAL and config.images_additional:
        return config.images_additional[0]
    return None


def process(in_path, out_path, config: MappingConfig, transform) -> ProcessResult:
    in_path = str(in_path)
    out_path = str(out_path)

    if os.path.abspath(in_path) == os.path.abspath(out_path):
        raise ConfigError("--out must differ from --in (refusing to overwrite the input)")

    result = ProcessResult(in_path=in_path, out_path=out_path)

    # --- preflight ---
    risks = preflight(in_path)
    if risks:
        if config.policy.on_preflight_risk == "error":
            raise PreflightError(
                "preflight found lossy content; aborting under policy "
                "on_preflight_risk=error:\n  "
                + "\n  ".join(f"[{r.severity}] {r.sheet or '<wb>'}: {r.detail}" for r in risks),
                risks=risks,
            )
        result.preflight_warnings = risks

    # --- load (read; we never save back to in_path) ---
    keep_vba = in_path.lower().endswith(".xlsm")
    wb = load_workbook(in_path, data_only=False, keep_links=True, keep_vba=keep_vba)

    try:
        for ws in _processed_worksheets(wb, config):
            header_block = detect_header(ws, config)
            mapped = resolve_columns(ws, config, header_block)
            result.mapping_summary.append(_summarize_mapping(mapped))

            start = header_block.data_start_row
            max_row = ws.max_row or start - 1
            for r in range(start, max_row + 1):
                result.rows_seen += 1
                # TITLE
                if mapped.title is not None:
                    _apply_text(
                        ws, r, mapped.title.col_index, transform.transform_title,
                        ColumnRole.TITLE, result,
                    )
                # LISTING
                if mapped.listing is not None:
                    _apply_text(
                        ws, r, mapped.listing.col_index, transform.transform_listing,
                        ColumnRole.LISTING, result,
                    )
                # IMAGE columns
                for mc in mapped.images_main:
                    _apply_image(ws, r, mc, config, transform, ColumnRole.IMAGE_MAIN, result)
                for mc in mapped.images_additional:
                    _apply_image(ws, r, mc, config, transform, ColumnRole.IMAGE_ADDITIONAL, result)

        wb.save(out_path)
    finally:
        wb.close()

    return result


def _apply_text(ws, row, col, fn, role: ColumnRole, result: ProcessResult) -> None:
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        result.skipped_merged.append(CellRef(ws.title, row, col))
        return
    old = cell.value
    old_str = None if old is None else str(old)
    new = fn(old_str)
    if new == old:
        return
    cell.value = new  # value only; styles/number_format/border untouched
    result.changed_cells.append(CellRef(ws.title, row, col))
    result.changed_by_role[role.value] = result.changed_by_role.get(role.value, 0) + 1


def _apply_image(ws, row, mc, config: MappingConfig, transform, role: ColumnRole,
                 result: ProcessResult) -> None:
    cell = ws.cell(row=row, column=mc.col_index)
    if isinstance(cell, MergedCell):
        result.skipped_merged.append(CellRef(ws.title, row, mc.col_index))
        return
    spec = _field_for_role(config, role)
    ic = images_mod.parse(cell, field_spec=spec, ws=ws)

    if ic.form is ImageForm.EMBEDDED_IMAGE:
        result.skipped_embedded.append(ic.ref)
        return

    new_urls = transform.transform_images(list(ic.urls))
    forced = mc.image_form_hint  # per-column override of detected form (for EMPTY cells)
    changed = images_mod.serialize(cell, ic, new_urls, forced_form=forced)
    if changed:
        result.changed_cells.append(CellRef(ws.title, row, mc.col_index))
        result.changed_by_role[role.value] = result.changed_by_role.get(role.value, 0) + 1


def _summarize_mapping(mapped) -> str:
    parts = [f"sheet={mapped.sheet!r}"]
    hb = mapped.header_block
    parts.append(f"header_rows={hb.header_rows} match_row={hb.match_row} data_start={hb.data_start_row}")
    if mapped.title:
        parts.append(f"title={mapped.title.column_letter}({mapped.title.matched_by})")
    if mapped.listing:
        parts.append(f"listing={mapped.listing.column_letter}({mapped.listing.matched_by})")
    if mapped.images_main:
        parts.append("images_main=" + ",".join(m.column_letter for m in mapped.images_main))
    if mapped.images_additional:
        parts.append("images_additional=" + ",".join(m.column_letter for m in mapped.images_additional))
    if mapped.key_column:
        parts.append(f"key={mapped.key_column.column_letter}(frozen)")
    parts.append(f"targets={sorted(mapped.target_col_indices)}")
    return " | ".join(parts)
