"""Shared fixtures + helpers for the test suite."""

from __future__ import annotations

import os

import pytest
from openpyxl import load_workbook

from ozon_excel_core.config import load_config
from ozon_excel_core.sample import gen_sample

CONFIG_PATH = os.path.join(
    os.path.dirname(os.path.dirname(__file__)), "fields.example.yaml"
)


@pytest.fixture
def config():
    return load_config(CONFIG_PATH)


@pytest.fixture
def config_path():
    return CONFIG_PATH


@pytest.fixture
def sample_xlsx(tmp_path):
    p = tmp_path / "sample.xlsx"
    gen_sample(p)
    return p


# --------------------------------------------------------------------------- #
# Helpers mirroring verifier signatures
# --------------------------------------------------------------------------- #


def _hyperlink_key(cell):
    hl = cell.hyperlink
    if hl is None:
        return None
    return (getattr(hl, "target", None), getattr(hl, "location", None))


def cell_signature(cell):
    return (cell.value, cell.number_format, _hyperlink_key(cell))


def diff_workbooks(path_a, path_b):
    """Return a dict {sheet: {(row,col): (sig_a, sig_b)}} of differing cells
    across the union of both workbooks' sheets/cells."""
    wb_a = load_workbook(str(path_a), data_only=False, keep_links=True)
    wb_b = load_workbook(str(path_b), data_only=False, keep_links=True)
    diffs = {}
    try:
        sheets = sorted(set(wb_a.sheetnames) | set(wb_b.sheetnames))
        for sheet in sheets:
            if sheet not in wb_a.sheetnames or sheet not in wb_b.sheetnames:
                diffs.setdefault(sheet, {})["__sheet__"] = ("present?", "present?")
                continue
            ws_a = wb_a[sheet]
            ws_b = wb_b[sheet]
            max_row = max(ws_a.max_row or 0, ws_b.max_row or 0)
            max_col = max(ws_a.max_column or 0, ws_b.max_column or 0)
            for r in range(1, max_row + 1):
                for c in range(1, max_col + 1):
                    sa = cell_signature(ws_a.cell(row=r, column=c))
                    sb = cell_signature(ws_b.cell(row=r, column=c))
                    if sa != sb:
                        diffs.setdefault(sheet, {})[(r, c)] = (sa, sb)
    finally:
        wb_a.close()
        wb_b.close()
    return diffs
