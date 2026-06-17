"""Multi-row header / data-start detection tests."""

from __future__ import annotations

import pytest
from openpyxl import Workbook, load_workbook

from ozon_excel_core.errors import MappingError
from ozon_excel_core.header import detect_header


def test_detects_multirow_header_and_data_start(sample_xlsx, config):
    wb = load_workbook(str(sample_xlsx))
    ws = wb["Шаблон для поставщика"]
    hb = detect_header(ws, config)
    assert hb.header_rows == [1, 2, 3]
    assert hb.match_row == 1
    assert hb.data_start_row == 4
    wb.close()


def test_no_keyword_match_raises(config):
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "totally"
    ws["B1"] = "unrelated"
    ws["A2"] = "data"
    with pytest.raises(MappingError):
        detect_header(ws, config)


def test_forced_match_row_and_data_start(sample_xlsx, config):
    from dataclasses import replace

    forced = replace(config, header=replace(config.header, match_row=1, data_start=4))
    wb = load_workbook(str(sample_xlsx))
    ws = wb["Шаблон для поставщика"]
    hb = detect_header(ws, forced)
    assert hb.match_row == 1
    assert hb.data_start_row == 4
    wb.close()
