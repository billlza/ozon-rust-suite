"""Every image-cell form: parse + serialize round-trip and re-emission."""

from __future__ import annotations

from openpyxl import Workbook

from ozon_excel_core import images as im
from ozon_excel_core.config import FieldSpec
from ozon_excel_core.model import ImageForm


def _cell(value=None, hyperlink=None):
    wb = Workbook()
    ws = wb.active
    c = ws["A1"]
    c.value = value
    if hyperlink is not None:
        c.hyperlink = hyperlink
    return wb, ws, c


def _multi_spec(read_delims):
    return FieldSpec(role="image_additional", read_delimiters=read_delims, write_delimiter="\n")


# --------------------------- PLAIN_URL --------------------------- #


def test_plain_url_parse_and_roundtrip():
    wb, ws, c = _cell("https://img.test/a.jpg")
    ic = im.parse(c, ws=ws)
    assert ic.form is ImageForm.PLAIN_URL
    assert ic.urls == ["https://img.test/a.jpg"]
    # no-op
    assert im.serialize(c, ic, ic.urls) is False
    assert c.value == "https://img.test/a.jpg"
    # change
    assert im.serialize(c, ic, ["https://cdn.test/a.jpg"]) is True
    assert c.value == "https://cdn.test/a.jpg"
    wb.close()


# --------------------------- HYPERLINK_FORMULA --------------------------- #


def test_hyperlink_formula_with_text():
    wb, ws, c = _cell('=HYPERLINK("https://img.test/m.jpg","фото")')
    ic = im.parse(c, ws=ws)
    assert ic.form is ImageForm.HYPERLINK_FORMULA
    assert ic.urls == ["https://img.test/m.jpg"]
    assert ic.display_text == "фото"
    assert im.serialize(c, ic, ic.urls) is False  # idempotent
    assert im.serialize(c, ic, ["https://cdn.test/m.jpg"]) is True
    assert c.value == '=HYPERLINK("https://cdn.test/m.jpg","фото")'
    wb.close()


def test_hyperlink_formula_no_text():
    wb, ws, c = _cell('=HYPERLINK("https://img.test/m.jpg")')
    ic = im.parse(c, ws=ws)
    assert ic.form is ImageForm.HYPERLINK_FORMULA
    assert ic.display_text is None
    assert im.serialize(c, ic, ic.urls) is False
    assert im.serialize(c, ic, ["https://cdn.test/x.jpg"]) is True
    assert c.value == '=HYPERLINK("https://cdn.test/x.jpg")'
    wb.close()


def test_hyperlink_formula_doubled_quote_escaping():
    # url has no quotes, display text contains a quote escaped as ""
    wb, ws, c = _cell('=HYPERLINK("https://img.test/m.jpg","say ""hi""")')
    ic = im.parse(c, ws=ws)
    assert ic.urls == ["https://img.test/m.jpg"]
    assert ic.display_text == 'say "hi"'
    # unchanged urls => byte-identical re-emit (quote escaping preserved)
    assert im.serialize(c, ic, ic.urls) is False
    assert c.value == '=HYPERLINK("https://img.test/m.jpg","say ""hi""")'
    wb.close()


# --------------------------- REAL_HYPERLINK --------------------------- #


def test_real_hyperlink_parse_and_change_url_only():
    wb, ws, c = _cell("Открыть фото", hyperlink="https://img.test/r.jpg")
    ic = im.parse(c, ws=ws)
    assert ic.form is ImageForm.REAL_HYPERLINK
    assert ic.urls == ["https://img.test/r.jpg"]
    assert ic.display_text == "Открыть фото"
    # no-op when url unchanged
    assert im.serialize(c, ic, ic.urls) is False
    # change url only — display text (cell.value) untouched
    assert im.serialize(c, ic, ["https://cdn.test/r.jpg"]) is True
    assert c.hyperlink.target == "https://cdn.test/r.jpg"
    assert c.value == "Открыть фото"
    wb.close()


# --------------------------- MULTI_URL --------------------------- #


def test_multi_url_newline_delimiter_preserved():
    val = "https://a.test/1.jpg\nhttps://a.test/2.jpg\nhttps://a.test/3.jpg"
    wb, ws, c = _cell(val)
    ic = im.parse(c, field_spec=_multi_spec(["\n", ",", " "]), ws=ws)
    assert ic.form is ImageForm.MULTI_URL
    assert ic.delimiter == "\n"
    assert len(ic.urls) == 3
    assert im.serialize(c, ic, ic.urls) is False
    assert c.value == val
    new = ["https://cdn/1.jpg", "https://cdn/2.jpg", "https://cdn/3.jpg"]
    assert im.serialize(c, ic, new) is True
    assert c.value == "\n".join(new)
    wb.close()


def test_multi_url_comma_delimiter_preserved():
    val = "https://a.test/1.jpg,https://a.test/2.jpg"
    wb, ws, c = _cell(val)
    ic = im.parse(c, field_spec=_multi_spec(["\n", ",", " "]), ws=ws)
    assert ic.form is ImageForm.MULTI_URL
    assert ic.delimiter == ","
    assert im.serialize(c, ic, ic.urls) is False
    new = ["https://cdn/1.jpg", "https://cdn/2.jpg"]
    assert im.serialize(c, ic, new) is True
    assert c.value == ",".join(new)
    wb.close()


def test_multi_url_space_delimiter_preserved():
    val = "https://a.test/1.jpg https://a.test/2.jpg"
    wb, ws, c = _cell(val)
    ic = im.parse(c, field_spec=_multi_spec(["\n", ",", " "]), ws=ws)
    assert ic.form is ImageForm.MULTI_URL
    assert ic.delimiter == " "
    assert im.serialize(c, ic, ic.urls) is False
    new = ["https://cdn/1.jpg", "https://cdn/2.jpg"]
    assert im.serialize(c, ic, new) is True
    assert c.value == " ".join(new)
    wb.close()


# ------------------- MULTI_URL with interleaved label ------------------- #


def test_multi_url_mixed_label_extracted_and_label_preserved():
    """A cell with 2+ urls and a non-url label between them must extract the
    urls individually (not treat the whole string as one opaque token) and
    preserve the label in position on a transforming write."""
    val = "https://img.test/1.jpg, см. фото, https://img.test/2.jpg"
    wb, ws, c = _cell(val)
    ic = im.parse(c, field_spec=_multi_spec(["\n", ",", " "]), ws=ws)
    assert ic.form is ImageForm.MULTI_URL
    # exactly the two real urls are extracted, label is NOT a url token
    assert ic.urls == ["https://img.test/1.jpg", "https://img.test/2.jpg"]
    # identity is a byte-for-byte no-op
    assert im.serialize(c, ic, ic.urls) is False
    assert c.value == val
    # a real transform rewrites both urls but keeps the label in its slot
    new = ["https://cdn/1.jpg", "https://cdn/2.jpg"]
    assert im.serialize(c, ic, new) is True
    assert c.value == "https://cdn/1.jpg,см. фото,https://cdn/2.jpg"
    wb.close()


# --------------- idempotency / byte-for-byte preservation --------------- #


def test_plain_url_surrounding_whitespace_preserved_on_identity():
    val = "   https://img.test/a.jpg  "
    wb, ws, c = _cell(val)
    ic = im.parse(c, ws=ws)
    assert ic.form is ImageForm.PLAIN_URL
    # transform sees the clean url
    assert ic.urls == ["https://img.test/a.jpg"]
    # but an identity write preserves the original value byte-for-byte
    assert im.serialize(c, ic, ic.urls) is False
    assert c.value == val
    wb.close()


def test_hyperlink_formula_internal_spacing_preserved_on_identity():
    val = '=HYPERLINK( "https://img.test/m.jpg" , "lbl" )'
    wb, ws, c = _cell(val)
    ic = im.parse(c, ws=ws)
    assert ic.form is ImageForm.HYPERLINK_FORMULA
    assert ic.urls == ["https://img.test/m.jpg"]
    assert im.serialize(c, ic, ic.urls) is False
    assert c.value == val  # not reformatted to canonical spacing
    wb.close()


def test_multi_url_delimiter_space_preserved_on_identity():
    val = "https://a.test/1.jpg, https://a.test/2.jpg"
    wb, ws, c = _cell(val)
    ic = im.parse(c, field_spec=_multi_spec(["\n", ",", " "]), ws=ws)
    assert ic.form is ImageForm.MULTI_URL
    assert im.serialize(c, ic, ic.urls) is False
    assert c.value == val  # ', ' not collapsed to ','
    wb.close()


# --------------------------- EMPTY --------------------------- #


def test_empty_cell():
    wb, ws, c = _cell(None)
    ic = im.parse(c, ws=ws)
    assert ic.form is ImageForm.EMPTY
    assert ic.urls == []
    # no forced form => no write
    assert im.serialize(c, ic, ["https://x"]) is False
    assert c.value is None
    # forced form writes
    assert im.serialize(c, ic, ["https://x.jpg"], forced_form=ImageForm.PLAIN_URL) is True
    assert c.value == "https://x.jpg"
    wb.close()


# --------------------------- EMBEDDED_IMAGE --------------------------- #


def test_embedded_image_detected_and_skipped():
    # Construct an anchored-image stand-in without Pillow: detection only reads
    # the anchor's _from.col / _from.row (0-based). This mirrors what openpyxl
    # exposes in ws._images for a real embedded raster image.
    from openpyxl.drawing.spreadsheet_drawing import (
        OneCellAnchor,
        AnchorMarker,
    )

    class _FakeImage:
        def __init__(self, col0, row0):
            self.anchor = OneCellAnchor(_from=AnchorMarker(col=col0, row=row0))

    wb = Workbook()
    ws = wb.active
    ws._images.append(_FakeImage(col0=0, row0=0))  # anchored to A1
    c = ws["A1"]
    ic = im.parse(c, ws=ws)
    assert ic.form is ImageForm.EMBEDDED_IMAGE
    # serialize never writes embedded images
    assert im.serialize(c, ic, ["https://x"]) is False
    # a cell that is NOT under an image is unaffected
    c2 = ws["B1"]
    c2.value = "https://plain.test/x.jpg"
    ic2 = im.parse(c2, ws=ws)
    assert ic2.form is ImageForm.PLAIN_URL
    wb.close()
