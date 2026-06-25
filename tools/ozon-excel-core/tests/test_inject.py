"""inject subcommand + inject_rows: rows land in mapped cells, verify stays OK."""

from __future__ import annotations

import json

from openpyxl import load_workbook

from ozon_excel_core.cli import main
from ozon_excel_core.inject import inject_rows


def test_inject_rows_lands_in_mapped_cells(tmp_path, config):
    from ozon_excel_core.sample import gen_sample

    sample = tmp_path / "sample.xlsx"
    out = tmp_path / "injected.xlsx"
    gen_sample(sample)

    rows = [
        {
            "title": "Новый заголовок",
            "listing": "Новое описание товара.",
            "primary_image": "https://cdn.test/main.jpg",
            "additional_images": [
                "https://cdn.test/a1.jpg",
                "https://cdn.test/a2.jpg",
            ],
        }
    ]
    placed = inject_rows(sample, out, config, rows)

    wb = load_workbook(str(out))
    ws = wb["Шаблон для поставщика"]
    r = placed["data_start_row"]
    assert ws[f"{placed['title_col']}{r}"].value == "Новый заголовок"
    assert ws[f"{placed['listing_col']}{r}"].value == "Новое описание товара."
    assert ws[f"{placed['main_image_col']}{r}"].value == "https://cdn.test/main.jpg"
    wb.close()


def test_inject_then_process_verify_ok(tmp_path, config_path, capsys):
    sample = tmp_path / "sample.xlsx"
    injected = tmp_path / "injected.xlsx"
    out = tmp_path / "out.xlsx"
    rows_json = tmp_path / "rows.json"

    assert main(["gen-sample", "--out", str(sample)]) == 0

    rows = [
        {
            "title": "Чайник электрический",
            "listing": "Объём 1.7 л.",
            "primary_image": "https://cdn.test/1001/main.jpg",
            "additional_images": ["https://cdn.test/1001/a1.jpg"],
        }
    ]
    rows_json.write_text(json.dumps(rows), encoding="utf-8")

    rc = main(["inject", "--in", str(sample), "--rows", str(rows_json),
               "--out", str(injected), "--config", config_path, "--quiet"])
    assert rc == 0
    assert injected.exists()

    # process --verify proves only mapped title/listing/image cells changed.
    rc = main(["process", "--in", str(injected), "--out", str(out),
               "--config", config_path, "--transform", "identity",
               "--verify", "--quiet"])
    assert rc == 0


def test_inject_in_equals_out_rejected(tmp_path, config_path):
    sample = tmp_path / "sample.xlsx"
    rows_json = tmp_path / "rows.json"
    main(["gen-sample", "--out", str(sample)])
    rows_json.write_text("[]", encoding="utf-8")
    rc = main(["inject", "--in", str(sample), "--rows", str(rows_json),
               "--out", str(sample), "--config", config_path])
    assert rc == 2
