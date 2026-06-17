"""Keyword (RU/EN/ZH) + letter resolution, multi-image columns, freeze set."""

from __future__ import annotations

import pytest
import yaml
from openpyxl import Workbook, load_workbook

from ozon_excel_core.config import parse_config
from ozon_excel_core.errors import MappingError
from ozon_excel_core.header import detect_header
from ozon_excel_core.mapper import resolve_columns


def _resolve(ws, config):
    hb = detect_header(ws, config)
    return resolve_columns(ws, config, hb)


def test_ru_keyword_resolution(sample_xlsx, config):
    wb = load_workbook(str(sample_xlsx))
    ws = wb["Шаблон для поставщика"]
    mapped = _resolve(ws, config)
    assert mapped.title.column_letter == "B"
    assert mapped.title.matched_by == "keyword"
    assert mapped.listing.column_letter == "C"
    assert len(mapped.images_main) == 1
    assert len(mapped.images_additional) == 2  # multi-url col + plain photo 2
    wb.close()


def test_freeze_set_excludes_physical_columns(sample_xlsx, config):
    wb = load_workbook(str(sample_xlsx))
    ws = wb["Шаблон для поставщика"]
    mapped = _resolve(ws, config)
    # targets are B,C,D,E,F => indices 2..6; key column A(1) NOT a target.
    assert mapped.target_col_indices == frozenset({2, 3, 4, 5, 6})
    # price (G=7), all the physical columns, are NOT targets => frozen.
    assert not mapped.is_target(7)
    assert not mapped.is_target(1)  # SKU/key column frozen
    wb.close()


def _build_header_only_wb(labels):
    wb = Workbook()
    ws = wb.active
    for ci, label in enumerate(labels, start=1):
        ws.cell(row=1, column=ci, value=label)
        ws.cell(row=2, column=ci, value=f"key{ci}")
    # a data row so data_start auto works
    ws.cell(row=3, column=1, value="SKU-X")
    return wb, ws


def test_en_header_variant(config):
    cfg = config
    wb, ws = _build_header_only_wb(
        ["SKU", "Title", "Description", "Main photo", "Additional photos", "Photo 2"]
    )
    # force data_start since our mini-sheet is small; auto should still work
    from dataclasses import replace

    cfg = replace(config, header=replace(config.header, scan_rows=2, data_start=3))
    mapped = _resolve(ws, cfg)
    assert mapped.title.column_letter == "B"
    assert mapped.listing.column_letter == "C"
    assert mapped.images_main and mapped.images_main[0].column_letter == "D"
    wb.close()


def test_zh_header_variant(config):
    from dataclasses import replace

    wb, ws = _build_header_only_wb(
        ["货号", "商品名称", "商品描述", "主图", "附加图片", "图片2"]
    )
    cfg = replace(config, header=replace(config.header, scan_rows=2, data_start=3))
    mapped = _resolve(ws, cfg)
    assert mapped.title.column_letter == "B"
    assert mapped.listing.column_letter == "C"
    assert mapped.images_main[0].column_letter == "D"
    wb.close()


def test_letter_pin_overrides_keyword(sample_xlsx):
    raw = {
        "version": 1,
        "header": {"scan_rows": 3, "match_row": 1, "data_start": 4},
        "match": {},
        "columns": {
            "title": {"letter": "C", "keywords": {"ru": ["Название товара"]}},
        },
        "policy": {},
    }
    cfg = parse_config(raw)
    wb = load_workbook(str(sample_xlsx))
    ws = wb["Шаблон для поставщика"]
    mapped = _resolve(ws, cfg)
    # letter pin C wins even though keyword "Название товара" is in B
    assert mapped.title.column_letter == "C"
    assert mapped.title.matched_by == "letter"
    wb.close()


def test_missing_required_raises(sample_xlsx):
    raw = {
        "version": 1,
        "header": {"scan_rows": 3, "match_row": 1, "data_start": 4},
        "match": {},
        "columns": {
            "title": {"keywords": {"ru": ["НесуществующаяКолонка"]}, "required": True},
        },
        "policy": {"on_missing_required": "error"},
    }
    cfg = parse_config(raw)
    wb = load_workbook(str(sample_xlsx))
    ws = wb["Шаблон для поставщика"]
    with pytest.raises(MappingError):
        _resolve(ws, cfg)
    wb.close()


def test_ambiguous_title_raises():
    raw = {
        "version": 1,
        "header": {"scan_rows": 2, "match_row": 1, "data_start": 3},
        "match": {},
        "columns": {
            "title": {"keywords": {"ru": ["фото"]}},  # matches two columns
        },
        "policy": {},
    }
    cfg = parse_config(raw)
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Главное фото"
    ws["B1"] = "Дополнительное фото"
    ws["A2"] = "x"
    with pytest.raises(MappingError):
        _resolve(ws, cfg)
    wb.close()


def test_image_keyword_overmatch_fails_closed():
    """A single image keyword spec that substring-matches a frozen physical
    column (e.g. 'Главное фото для внутреннего архива' ⊃ 'Главное фото') must
    fail-closed with MappingError, exactly like the single-valued title/listing
    fields — never silently add the frozen column to the write set."""
    raw = {
        "version": 1,
        "header": {"scan_rows": 2, "match_row": 1, "data_start": 3},
        "match": {"mode": "contains"},
        "columns": {
            "images_main": {"keywords": {"ru": ["Ссылка на главное фото", "Главное фото"]}},
        },
        "policy": {},
    }
    cfg = parse_config(raw)
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Артикул"
    ws["D1"] = "Ссылка на главное фото"          # legit target
    ws["G1"] = "Главное фото для внутреннего архива"  # frozen, substring-contains keyword
    ws["A2"] = "k"
    ws["A3"] = "SKU-1"
    with pytest.raises(MappingError):
        _resolve(ws, cfg)
    wb.close()


def test_image_multi_spec_distinct_columns_ok():
    """The legitimate multi-image design (one keyword spec per distinct column)
    still resolves to multiple columns — the over-match guard only refuses a
    SINGLE spec landing on >1 column."""
    raw = {
        "version": 1,
        "header": {"scan_rows": 2, "match_row": 1, "data_start": 3},
        "match": {"mode": "contains"},
        "columns": {
            "images_main": {"keywords": {"ru": ["Главное фото"]}},
            "images_additional": [
                {"keywords": {"ru": ["Дополнительные фото"]}, "form": "multi_url"},
                {"keywords": {"ru": ["Ссылка на фото 2"]}, "form": "plain_url"},
            ],
        },
        "policy": {},
    }
    cfg = parse_config(raw)
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Артикул"
    ws["D1"] = "Ссылка на главное фото"
    ws["E1"] = "Дополнительные фото"
    ws["F1"] = "Ссылка на фото 2"
    ws["A2"] = "k"
    ws["A3"] = "SKU-1"
    mapped = _resolve(ws, cfg)
    assert [m.column_letter for m in mapped.images_main] == ["D"]
    assert [m.column_letter for m in mapped.images_additional] == ["E", "F"]
    assert sorted(mapped.target_col_indices) == [4, 5, 6]
    wb.close()


def test_contradictory_letter_pin_raises():
    raw = {
        "version": 1,
        "header": {"scan_rows": 2, "match_row": 1, "data_start": 3},
        "match": {},
        "columns": {
            "title": {"letter": "B"},
            "listing": {"letter": "B"},
        },
        "policy": {},
    }
    cfg = parse_config(raw)
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "x"
    ws["B1"] = "y"
    ws["A2"] = "d"
    with pytest.raises(MappingError):
        _resolve(ws, cfg)
    wb.close()
