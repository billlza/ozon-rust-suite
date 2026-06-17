"""No-op transform => identical output at the cell-signature level."""

from __future__ import annotations

from openpyxl import load_workbook

from ozon_excel_core import process, verify
from ozon_excel_core.transforms.identity import IdentityTransform

from conftest import diff_workbooks


def test_identity_produces_zero_changes(sample_xlsx, config, tmp_path):
    out = tmp_path / "out_identity.xlsx"
    result = process(sample_xlsx, out, config, IdentityTransform())
    assert result.total_changed() == 0

    diffs = diff_workbooks(sample_xlsx, out)
    assert diffs == {}, f"identity must not change any cell, got: {diffs}"

    report = verify(sample_xlsx, out, config)
    assert report.ok is True
    assert report.summary["expected_changes"] == 0
    assert report.summary["unexpected_changes"] == 0


def test_identity_preserves_image_cell_bytes(sample_xlsx, config, tmp_path):
    out = tmp_path / "out_identity.xlsx"
    process(sample_xlsx, out, config, IdentityTransform())

    wb_a = load_workbook(str(sample_xlsx))
    wb_b = load_workbook(str(out))
    ws_a = wb_a["Шаблон для поставщика"]
    ws_b = wb_b["Шаблон для поставщика"]
    # main photo HYPERLINK formula (D), additional multi-url (E), plain photo (F)
    for coord in ("D4", "E4", "F4", "D5", "E5", "D6", "E6"):
        assert ws_a[coord].value == ws_b[coord].value, coord
    wb_a.close()
    wb_b.close()
