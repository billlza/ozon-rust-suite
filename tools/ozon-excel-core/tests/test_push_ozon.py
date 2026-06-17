"""push_ozon.py: fully MOCKED — no real network, no writes.

A FakeOzonClient records every call and returns canned product_info /
attributes. Covers:
  - dry-run prints a plan and makes ZERO write calls;
  - --apply calls pictures/import with images[0] == the Excel's main-image url
    and the right product_id;
  - --keep-existing appends the product's current images after the new ones;
  - --push-title builds an import payload changing ONLY the name (other
    attributes preserved);
  - --limit is honored;
  - the standalone CLI subcommand wiring works against the fake.
"""

from __future__ import annotations

import importlib.util
import os
import sys

import pytest


# --------------------------------------------------------------------------- #
# Load the standalone script the same way the suite loads demo_real.py.
# --------------------------------------------------------------------------- #
def _load_push():
    here = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    path = os.path.join(here, "scripts", "push_ozon.py")
    spec = importlib.util.spec_from_file_location("push_ozon", path)
    mod = importlib.util.module_from_spec(spec)
    sys.modules["push_ozon"] = mod
    spec.loader.exec_module(mod)
    return mod


push = _load_push()


# --------------------------------------------------------------------------- #
# Fake Ozon client: records calls, returns canned data, performs NO network.
# --------------------------------------------------------------------------- #
class FakeOzonClient:
    def __init__(self, products):
        # products: {offer_id: {"product_id", "name", "primary_image", "images",
        #            "attributes"(optional dict)}}
        self.products = products
        self.calls = []  # list of (method, payload)
        self._next_primary = {}  # offer_id -> primary to report after a write

    # ---- reads ---- #
    def product_info_by_offer(self, offer_ids):
        self.calls.append(("product_info_by_offer", list(offer_ids)))
        out = []
        for oid in offer_ids:
            p = self.products.get(str(oid))
            if p is None:
                continue
            primary = self._next_primary.get(str(oid), p["primary_image"])
            out.append({
                "id": p["product_id"],
                "offer_id": str(oid),
                "name": p["name"],
                "primary_image": primary,
                "images": list(p["images"]),
            })
        return out

    def product_attributes(self, product_id, visibility="ALL"):
        self.calls.append(("product_attributes", product_id))
        for p in self.products.values():
            if p["product_id"] == product_id:
                return dict(p.get("attributes") or {})
        return {}

    # ---- writes (must NOT be called in dry-run) ---- #
    def pictures_import(self, product_id, images, **kw):
        self.calls.append(("pictures_import", {"product_id": product_id, "images": list(images)}))
        # Reflect: images[0] becomes the new primary on the next info fetch.
        for oid, p in self.products.items():
            if p["product_id"] == product_id and images:
                self._next_primary[oid] = images[0]
        return {"pictures": [
            {"url": u, "state": "imported", "is_primary": (i == 0)}
            for i, u in enumerate(images)
        ]}

    def product_import(self, items):
        self.calls.append(("product_import", items))
        return 999001

    def product_import_info(self, task_id):
        self.calls.append(("product_import_info", task_id))
        return {"status": "imported"}

    # ---- helpers ---- #
    def write_calls(self):
        return [c for c in self.calls if c[0] in (
            "pictures_import", "product_import", "product_import_info")]


def _canned_products():
    return {
        "SKU-1001": {
            "product_id": 1001,
            "name": "old name 1",
            "primary_image": "https://old.test/1001/old-main.jpg",
            "images": ["https://old.test/1001/old-main.jpg",
                       "https://old.test/1001/old-extra.jpg"],
            "attributes": {
                "offer_id": "SKU-1001",
                "category_id": 17028922,
                "name": "old name 1",
                "barcode": "4600123456789",
                "attributes": [
                    {"id": 85, "values": [{"dictionary_value_id": 1, "value": "BrandX"}]},
                    {"id": 4180, "values": [{"value": "old name 1"}]},
                ],
            },
        },
        "SKU-1002": {
            "product_id": 1002,
            "name": "old name 2",
            "primary_image": "https://old.test/1002/old-main.jpg",
            "images": ["https://old.test/1002/old-main.jpg"],
            "attributes": {
                "offer_id": "SKU-1002",
                "category_id": 17028922,
                "name": "old name 2",
                "attributes": [{"id": 85, "values": [{"value": "BrandY"}]}],
            },
        },
    }


# --------------------------------------------------------------------------- #
# A relisted workbook fixture: the core sample with image/title cells rewritten.
# --------------------------------------------------------------------------- #
@pytest.fixture
def relisted_xlsx(tmp_path, config):
    """Build a sample workbook then overwrite the first two data rows' title +
    image cells with NEW (relisted) values. READ side under test only reads it."""
    from openpyxl import load_workbook

    from ozon_excel_core.header import detect_header
    from ozon_excel_core.mapper import resolve_columns
    from ozon_excel_core.sample import gen_sample

    p = tmp_path / "relisted.xlsx"
    gen_sample(p)

    wb = load_workbook(str(p))
    ws = wb["Шаблон для поставщика"]
    hb = detect_header(ws, config)
    mapped = resolve_columns(ws, config, hb)
    ds = hb.data_start_row

    title_col = mapped.title.col_index
    main_col = mapped.images_main[0].col_index          # plain url (we overwrite)
    add_cols = [mc.col_index for mc in mapped.images_additional]

    # Row 1 (SKU-1001): new title, new main image (plain url) + a multi-url add.
    ws.cell(row=ds, column=title_col, value="NEW relisted name 1")
    ws.cell(row=ds, column=main_col, value="https://new.test/1001/relisted-main.jpg")
    ws.cell(row=ds, column=add_cols[0],
            value="https://new.test/1001/relisted-a1.jpg\nhttps://new.test/1001/relisted-a2.jpg")
    if len(add_cols) > 1:
        ws.cell(row=ds, column=add_cols[1], value="https://new.test/1001/relisted-p2.jpg")

    # Row 2 (SKU-1002): new title + single new main image.
    ws.cell(row=ds + 1, column=title_col, value="NEW relisted name 2")
    ws.cell(row=ds + 1, column=main_col, value="https://new.test/1002/relisted-main.jpg")

    wb.save(str(p))
    wb.close()
    return p


# --------------------------------------------------------------------------- #
# read_rows: offer_id + ordered image urls + title come out of the xlsx.
# --------------------------------------------------------------------------- #
def test_read_rows_collects_offer_images_title(relisted_xlsx, config):
    rows, layout = push.read_rows(relisted_xlsx, config)
    assert layout["offer_col"] == "A"  # Артикул is the key column
    by_offer = {r.offer_id: r for r in rows}

    r1 = by_offer["SKU-1001"]
    assert r1.title == "NEW relisted name 1"
    # main image first (primary), then the multi-url additionals in order, then photo-2
    assert r1.image_urls[0] == "https://new.test/1001/relisted-main.jpg"
    assert r1.image_urls[:3] == [
        "https://new.test/1001/relisted-main.jpg",
        "https://new.test/1001/relisted-a1.jpg",
        "https://new.test/1001/relisted-a2.jpg",
    ]
    assert "https://new.test/1001/relisted-p2.jpg" in r1.image_urls


def test_offer_col_letter_and_keyword(relisted_xlsx, config):
    # explicit column LETTER
    rows_a, layout_a = push.read_rows(relisted_xlsx, config, offer_col_arg="A")
    assert layout_a["offer_col"] == "A"
    assert {r.offer_id for r in rows_a} >= {"SKU-1001", "SKU-1002"}
    # header KEYWORD ("Артикул") resolves to the same column A
    rows_k, layout_k = push.read_rows(relisted_xlsx, config, offer_col_arg="Артикул")
    assert layout_k["offer_col"] == "A"


def _capture():
    lines = []
    return lines, lambda *a: lines.append(" ".join(str(x) for x in a))


# --------------------------------------------------------------------------- #
# 1. dry-run: prints a plan, ZERO write calls.
# --------------------------------------------------------------------------- #
def test_dry_run_plans_and_makes_no_writes(relisted_xlsx, config):
    rows, _ = push.read_rows(relisted_xlsx, config)
    client = FakeOzonClient(_canned_products())
    lines, out = _capture()

    plans, results, confs = push.run_push(
        client, rows, apply=False, limit=1,
        keep_existing=False, push_title=False, out=out,
    )

    text = "\n".join(lines)
    assert "DRY-RUN" in text
    assert "PRIMARY" in text
    assert "https://new.test/1001/relisted-main.jpg" in text
    # NO write calls whatsoever.
    assert client.write_calls() == []
    # resolution read happened.
    assert ("product_info_by_offer", ["SKU-1001"]) in client.calls
    assert results == [] and confs == []
    # plan primary is the excel main image.
    assert plans[0]["primary"] == "https://new.test/1001/relisted-main.jpg"
    assert plans[0]["product_id"] == 1001


# --------------------------------------------------------------------------- #
# 2. --apply: pictures/import with images[0] == excel main url + right product_id.
# --------------------------------------------------------------------------- #
def test_apply_pushes_images_primary_first(relisted_xlsx, config):
    rows, _ = push.read_rows(relisted_xlsx, config)
    client = FakeOzonClient(_canned_products())
    lines, out = _capture()

    push.run_push(
        client, rows, apply=True, limit=1,
        keep_existing=False, push_title=False, out=out,
    )

    pic_calls = [c for c in client.calls if c[0] == "pictures_import"]
    assert len(pic_calls) == 1
    payload = pic_calls[0][1]
    assert payload["product_id"] == 1001
    assert payload["images"][0] == "https://new.test/1001/relisted-main.jpg"
    # no existing images appended (keep_existing was False)
    assert all(not u.startswith("https://old.test") for u in payload["images"])
    # title was NOT pushed
    assert not any(c[0] == "product_import" for c in client.calls)
    # post-apply confirmation re-fetched and matched the pushed primary.
    text = "\n".join(lines)
    assert "confirm SKU-1001" in text and "[OK]" in text


# --------------------------------------------------------------------------- #
# 3. --keep-existing: current images appended AFTER the new ones.
# --------------------------------------------------------------------------- #
def test_keep_existing_appends_current_images(relisted_xlsx, config):
    rows, _ = push.read_rows(relisted_xlsx, config)
    client = FakeOzonClient(_canned_products())
    _, out = _capture()

    push.run_push(
        client, rows, apply=True, limit=1,
        keep_existing=True, push_title=False, out=out,
    )

    payload = [c for c in client.calls if c[0] == "pictures_import"][0][1]
    imgs = payload["images"]
    # new ones first
    assert imgs[0] == "https://new.test/1001/relisted-main.jpg"
    # the product's current images appended after, de-duplicated, order preserved
    assert "https://old.test/1001/old-main.jpg" in imgs
    assert "https://old.test/1001/old-extra.jpg" in imgs
    assert imgs.index("https://new.test/1001/relisted-main.jpg") < imgs.index(
        "https://old.test/1001/old-main.jpg")


def test_keep_existing_unit_dedupes():
    out = push.build_image_list(
        ["a", "b"], ["b", "c"], keep_existing=True)
    assert out == ["a", "b", "c"]  # 'b' not duplicated
    # without keep_existing, only the new ones
    assert push.build_image_list(["a"], ["c"], keep_existing=False) == ["a"]


# --------------------------------------------------------------------------- #
# 4. --push-title: import payload changes ONLY the name, preserves other attrs.
# --------------------------------------------------------------------------- #
def test_push_title_changes_only_name(relisted_xlsx, config):
    rows, _ = push.read_rows(relisted_xlsx, config)
    client = FakeOzonClient(_canned_products())
    _, out = _capture()

    push.run_push(
        client, rows, apply=True, limit=1,
        keep_existing=False, push_title=True, out=out,
    )

    imp = [c for c in client.calls if c[0] == "product_import"]
    assert len(imp) == 1
    item = imp[0][1][0]
    original = _canned_products()["SKU-1001"]["attributes"]

    # name changed to the excel title
    assert item["name"] == "NEW relisted name 1"
    # every non-name field preserved exactly
    assert item["offer_id"] == original["offer_id"]
    assert item["category_id"] == original["category_id"]
    assert item["barcode"] == original["barcode"]
    # the non-name attribute (id 85) is byte-identical
    attr85 = [a for a in item["attributes"] if a["id"] == 85]
    assert attr85 == [a for a in original["attributes"] if a["id"] == 85]
    # the name attribute (id 4180) was updated to the new title, nothing else
    attr_name = [a for a in item["attributes"] if a["id"] == 4180]
    assert attr_name == [{"id": 4180, "values": [{"value": "NEW relisted name 1"}]}]
    # a status poll happened
    assert any(c[0] == "product_import_info" for c in client.calls)


def test_push_title_skips_when_attributes_thin():
    # Missing category + attributes -> refuse, do not build an import item.
    item, reason = push.build_title_import_item({"offer_id": "X"}, "new")
    assert item is None
    assert "missing" in reason
    # empty attributes -> refuse
    item2, reason2 = push.build_title_import_item({}, "new")
    assert item2 is None


def test_push_title_skip_does_not_call_import(relisted_xlsx, config):
    # Strip the attributes so the title push is skipped per-product.
    products = _canned_products()
    products["SKU-1001"]["attributes"] = {"offer_id": "SKU-1001"}  # thin
    client = FakeOzonClient(products)
    rows, _ = push.read_rows(relisted_xlsx, config)
    lines, out = _capture()

    plans, _, _ = push.run_push(
        client, rows, apply=True, limit=1,
        keep_existing=False, push_title=True, out=out,
    )
    assert plans[0]["title_push"] is False
    assert plans[0]["title_skip_reason"] is not None
    assert not any(c[0] == "product_import" for c in client.calls)
    assert "title change SKIPPED" in "\n".join(lines)


# --------------------------------------------------------------------------- #
# 5. --limit is honored.
# --------------------------------------------------------------------------- #
def test_limit_caps_touched_products(relisted_xlsx, config):
    rows, _ = push.read_rows(relisted_xlsx, config)
    assert len(rows) >= 2
    client = FakeOzonClient(_canned_products())
    _, out = _capture()

    plans, _, _ = push.run_push(
        client, rows, apply=True, limit=1,
        keep_existing=False, push_title=False, out=out,
    )
    # only one product planned + only one pictures_import
    assert len(plans) == 1
    assert plans[0]["offer_id"] == "SKU-1001"
    assert len([c for c in client.calls if c[0] == "pictures_import"]) == 1

    # limit=2 touches both
    client2 = FakeOzonClient(_canned_products())
    plans2, _, _ = push.run_push(
        client2, rows, apply=True, limit=2,
        keep_existing=False, push_title=False, out=out,
    )
    assert len(plans2) == 2
    assert len([c for c in client2.calls if c[0] == "pictures_import"]) == 2


# --------------------------------------------------------------------------- #
# 6. unresolved offer_id is skipped cleanly (no writes for it).
# --------------------------------------------------------------------------- #
def test_unresolved_offer_is_skipped(relisted_xlsx, config):
    products = _canned_products()
    del products["SKU-1001"]  # first row can't resolve
    client = FakeOzonClient(products)
    rows, _ = push.read_rows(relisted_xlsx, config)
    lines, out = _capture()

    plans, results, _ = push.run_push(
        client, rows, apply=True, limit=1,
        keep_existing=False, push_title=False, out=out,
    )
    assert plans[0]["resolved"] is False
    assert results == []  # nothing applied
    assert client.write_calls() == []
    assert "SKIP" in "\n".join(lines)


# --------------------------------------------------------------------------- #
# 7. the OzonClient gained the verified push methods (shape only, no network).
# --------------------------------------------------------------------------- #
def test_ozon_client_push_methods_post_correct_bodies():
    from ozon_excel_core.ozon_api import OzonClient

    posted = []

    client = OzonClient("cid", "key")

    def fake_post(path, body):
        posted.append((path, body))
        if path == "/v1/product/pictures/import":
            return {"result": {"pictures": [{"url": body["images"][0],
                                             "state": "imported", "is_primary": True}]}}
        if path == "/v4/product/info/attributes":
            return {"result": {"items": [{"offer_id": "X", "category_id": 1}]}}
        if path == "/v1/product/import":
            return {"result": {"task_id": 7}}
        if path == "/v1/product/import/info":
            return {"result": {"status": "imported"}}
        if path == "/v3/product/info/list":
            return {"result": {"items": [{"id": 5, "offer_id": "X"}]}}
        return {}

    client._post = fake_post

    res = client.pictures_import(42, ["u1", "u2"])
    assert posted[-1] == ("/v1/product/pictures/import",
                          {"product_id": 42, "images": ["u1", "u2"],
                           "color_image": "", "images360": []})
    assert res["pictures"][0]["is_primary"] is True

    attrs = client.product_attributes(42)
    assert posted[-1][0] == "/v4/product/info/attributes"
    assert posted[-1][1] == {"filter": {"product_id": [42], "visibility": "ALL"}, "limit": 1}
    assert attrs["offer_id"] == "X"

    tid = client.product_import([{"offer_id": "X"}])
    assert tid == 7 and posted[-1][0] == "/v1/product/import"

    info = client.product_import_info(7)
    assert info["status"] == "imported"

    items = client.product_info_by_offer(["X"])
    assert posted[-1] == ("/v3/product/info/list", {"offer_id": ["X"]})
    assert items[0]["offer_id"] == "X"


# --------------------------------------------------------------------------- #
# 8. CLI subcommand wiring: `push-ozon` dry-run runs against a fake client.
# --------------------------------------------------------------------------- #
def test_cli_subcommand_dry_run(relisted_xlsx, config_path, monkeypatch, capsys):
    import ozon_excel_core.cli as cli

    fake = FakeOzonClient(_canned_products())
    # Patch OzonClient.from_env (imported inside push_ozon.main) to the fake.
    import ozon_excel_core.ozon_api as ozon_api
    monkeypatch.setattr(ozon_api.OzonClient, "from_env",
                        classmethod(lambda cls, **kw: fake))

    rc = cli.main(["push-ozon", "--in", str(relisted_xlsx),
                   "--config", config_path, "--limit", "1"])
    assert rc == 0
    captured = capsys.readouterr().out
    assert "DRY-RUN" in captured
    assert client_made_no_writes(fake)


def client_made_no_writes(fake):
    return fake.write_calls() == []
