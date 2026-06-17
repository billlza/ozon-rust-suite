"""RelistTransform + demo_real.py --mock: fully offline, no network, no spend.

Covers:
  - transform_images maps http(s) urls -> hosted restyle urls, passes non-urls
    through unchanged;
  - on-disk caching: a second transform_images on the same url is served from
    cache and does NOT call the imagegen fn again (call count asserted);
  - the relist transform, run through the core writer, changes ONLY content
    cells (title/listing/image) and verify reports the frozen set byte-identical;
  - demo_real.py --mock runs end-to-end and verify reports ONLY title/listing/
    image roles changed, with every other (~45) frozen cell identical.
"""

from __future__ import annotations

import importlib.util
import os
import sys

import pytest

from ozon_excel_core import process, verify
from ozon_excel_core.header import detect_header
from ozon_excel_core.mapper import resolve_columns
from ozon_excel_core.transforms.relist import DEFAULT_PROMPT, RelistTransform

from openpyxl import load_workbook

from conftest import diff_workbooks

_TINY_PNG = b"\x89PNG\r\n\x1a\n" + b"\x00" * 16


# --------------------------------------------------------------------------- #
# Fakes (no network, no spend)
# --------------------------------------------------------------------------- #
class FakeImageGen:
    def __init__(self):
        self.calls = 0
        self.seen = []

    def __call__(self, image_bytes, prompt, **kwargs):
        self.calls += 1
        self.seen.append((prompt, kwargs))
        return _TINY_PNG


class FakeHost:
    name = "fake"

    def __init__(self):
        self.store = {}
        self.put_calls = 0

    def put(self, filename, data):
        self.put_calls += 1
        url = f"https://fake-host.test/{filename}"
        self.store[url] = data
        return url


@pytest.fixture
def fake_source_get(monkeypatch):
    """Make the transform's source-image download offline."""
    import ozon_excel_core.transforms.relist as relist_mod

    class _Resp:
        status_code = 200
        content = b"source-image-bytes"

    monkeypatch.setattr(relist_mod.requests, "get", lambda url, **kw: _Resp())
    return relist_mod


def _make_transform(tmp_path, imagegen=None, host=None):
    return RelistTransform(
        imagegen_fn=imagegen or FakeImageGen(),
        host=host or FakeHost(),
        image_api_base="https://fake.test",
        image_api_key="fake-key",
        image_model="gpt-image-2-vip",
        cache_dir=str(tmp_path / "cache"),
    )


# --------------------------------------------------------------------------- #
# 1. mapping + non-url passthrough
# --------------------------------------------------------------------------- #
def test_transform_images_maps_urls_and_passes_non_urls(tmp_path, fake_source_get):
    gen = FakeImageGen()
    host = FakeHost()
    t = _make_transform(tmp_path, imagegen=gen, host=host)

    urls = [
        "https://img.ozon.test/1/main.jpg",
        "см. фото",  # non-url label -> passthrough
        "https://img.ozon.test/2/a1.jpg",
        "",  # empty -> passthrough
    ]
    out = t.transform_images(urls)

    # non-urls unchanged, in place
    assert out[1] == "см. фото"
    assert out[3] == ""
    # urls rewritten to the fake host
    assert out[0].startswith("https://fake-host.test/")
    assert out[2].startswith("https://fake-host.test/")
    assert out[0] != out[2]
    # one generation + one host.put per distinct url
    assert gen.calls == 2
    assert host.put_calls == 2
    # the configured prompt/model were passed through
    assert gen.seen[0][0] == DEFAULT_PROMPT
    assert gen.seen[0][1]["model"] == "gpt-image-2-vip"


# --------------------------------------------------------------------------- #
# 2. cache: a second call on the same url does NOT re-generate
# --------------------------------------------------------------------------- #
def test_second_call_served_from_cache(tmp_path, fake_source_get):
    gen = FakeImageGen()
    host = FakeHost()
    t = _make_transform(tmp_path, imagegen=gen, host=host)

    url = "https://img.ozon.test/1/main.jpg"
    first = t.transform_images([url])
    assert gen.calls == 1
    assert host.put_calls == 1

    second = t.transform_images([url])
    # identical result, but NO new generation / host upload
    assert second == first
    assert gen.calls == 1, "imagegen must not be called again for a cached url"
    assert host.put_calls == 1, "host must not be called again for a cached url"

    # A fresh transform sharing the same cache_dir also hits the cache.
    t2 = RelistTransform(
        imagegen_fn=gen,
        host=host,
        image_api_base="https://fake.test",
        image_api_key="fake-key",
        cache_dir=str(tmp_path / "cache"),
    )
    third = t2.transform_images([url])
    assert third == first
    assert gen.calls == 1


def test_cache_key_depends_on_model_and_prompt(tmp_path, fake_source_get):
    gen = FakeImageGen()
    host = FakeHost()
    url = "https://img.ozon.test/1/main.jpg"

    t1 = _make_transform(tmp_path, imagegen=gen, host=host)
    t1.transform_images([url])
    assert gen.calls == 1

    # Different prompt -> different cache key -> regenerate.
    t2 = RelistTransform(
        imagegen_fn=gen,
        host=host,
        image_api_base="https://fake.test",
        image_api_key="fake-key",
        prompt="a different prompt entirely",
        cache_dir=str(tmp_path / "cache"),
    )
    t2.transform_images([url])
    assert gen.calls == 2


# --------------------------------------------------------------------------- #
# 3. through the core writer: only content cells change
# --------------------------------------------------------------------------- #
def test_relist_changes_only_content(sample_xlsx, config, tmp_path, fake_source_get):
    out = tmp_path / "relist_out.xlsx"
    t = _make_transform(tmp_path)

    result = process(sample_xlsx, out, config, t)
    assert result.total_changed() > 0

    # target columns + data start
    wb = load_workbook(str(sample_xlsx))
    ws = wb["Шаблон для поставщика"]
    hb = detect_header(ws, config)
    mapped = resolve_columns(ws, config, hb)
    wb.close()
    targets, data_start = mapped.target_col_indices, hb.data_start_row

    diffs = diff_workbooks(sample_xlsx, out)
    assert set(diffs) <= {"Шаблон для поставщика"}
    for (row, col), _ in diffs.get("Шаблон для поставщика", {}).items():
        assert col in targets, f"non-target column {col} changed at row {row}"
        assert row >= data_start, f"header/structural row {row} changed"

    report = verify(sample_xlsx, out, config)
    assert report.ok is True
    assert report.unexpected_changes == []
    assert len(report.expected_changes) > 0
    assert {d.role for d in report.expected_changes} <= {"title", "listing", "image"}


# --------------------------------------------------------------------------- #
# 4. demo_real.py --mock end-to-end
# --------------------------------------------------------------------------- #
def _load_demo():
    here = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    path = os.path.join(here, "scripts", "demo_real.py")
    spec = importlib.util.spec_from_file_location("demo_real", path)
    mod = importlib.util.module_from_spec(spec)
    sys.modules["demo_real"] = mod
    spec.loader.exec_module(mod)
    return mod


def test_demo_real_mock_end_to_end(tmp_path):
    demo = _load_demo()
    rc = demo.run(mock=True, products_n=3, out_dir=str(tmp_path / "demo"))
    assert rc == 0

    config_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "fields.example.yaml",
    )
    from ozon_excel_core import load_config

    config = load_config(config_path)
    real_in = tmp_path / "demo" / "real_input.xlsx"
    real_out = tmp_path / "demo" / "real_output.xlsx"
    assert real_in.exists() and real_out.exists()

    report = verify(real_in, real_out, config)
    assert report.ok is True
    assert report.unexpected_changes == []
    # ONLY content roles changed
    assert {d.role for d in report.expected_changes} <= {"title", "listing", "image"}
    # the bulk of the ~50-column sheet stayed frozen & byte-identical
    assert report.summary["frozen_cells_compared"] >= 45

    # cross-check: every diff between input and output is a mapped target cell
    wb = load_workbook(str(real_in))
    ws = wb["Шаблон для поставщика"]
    hb = detect_header(ws, config)
    mapped = resolve_columns(ws, config, hb)
    wb.close()
    diffs = diff_workbooks(real_in, real_out)
    assert set(diffs) <= {"Шаблон для поставщика"}
    for (row, col), _ in diffs.get("Шаблон для поставщика", {}).items():
        assert col in mapped.target_col_indices
        assert row >= hb.data_start_row
