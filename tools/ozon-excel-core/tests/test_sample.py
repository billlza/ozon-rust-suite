"""gen-sample produces a valid, mappable, multi-sheet workbook."""

from __future__ import annotations

from openpyxl import load_workbook

from ozon_excel_core.sample import gen_sample
from ozon_excel_core.header import detect_header


def test_sample_structure(tmp_path, config):
    p = tmp_path / "s.xlsx"
    gen_sample(p)
    wb = load_workbook(str(p))
    assert set(wb.sheetnames) == {
        "Шаблон для поставщика",
        "validation",
        "Инструкция",
    }
    # hidden validation sheet
    assert wb["validation"].sheet_state == "hidden"

    ws = wb["Шаблон для поставщика"]
    assert ws.max_column >= 48  # ~50 columns
    assert ws.max_column <= 52

    hb = detect_header(ws, config)
    assert hb.data_start_row == 4
    assert hb.header_rows == [1, 2, 3]
    wb.close()


def test_sample_has_both_image_forms(tmp_path):
    p = tmp_path / "s.xlsx"
    gen_sample(p)
    wb = load_workbook(str(p))
    ws = wb["Шаблон для поставщика"]
    # D = main photo HYPERLINK formula
    assert str(ws["D4"].value).upper().startswith("=HYPERLINK")
    # E = additional multi-url (contains a delimiter)
    e = str(ws["E4"].value)
    assert "\n" in e or "," in e or " " in e
    assert "http" in e
    # F = plain single url
    assert str(ws["F4"].value).startswith("http")
    wb.close()
