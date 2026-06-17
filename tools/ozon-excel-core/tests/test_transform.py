"""Real transform => only target cells differ; everything else identical."""

from __future__ import annotations

from ozon_excel_core import process, verify
from ozon_excel_core.header import detect_header
from ozon_excel_core.mapper import resolve_columns
from ozon_excel_core.transforms.example import ExampleAllTransform

from openpyxl import load_workbook

from conftest import diff_workbooks


def _target_cols_and_start(sample_xlsx, config):
    wb = load_workbook(str(sample_xlsx))
    ws = wb["Шаблон для поставщика"]
    hb = detect_header(ws, config)
    mapped = resolve_columns(ws, config, hb)
    wb.close()
    return mapped.target_col_indices, hb.data_start_row


def test_only_target_cells_change(sample_xlsx, config, tmp_path):
    out = tmp_path / "out_example.xlsx"
    result = process(sample_xlsx, out, config, ExampleAllTransform())
    assert result.total_changed() > 0

    targets, data_start = _target_cols_and_start(sample_xlsx, config)

    diffs = diff_workbooks(sample_xlsx, out)
    # only the main sheet should have diffs
    assert set(diffs) <= {"Шаблон для поставщика"}
    for (row, col), _ in diffs.get("Шаблон для поставщика", {}).items():
        assert col in targets, f"non-target column {col} changed at row {row}"
        assert row >= data_start, f"header/structural row {row} changed"


def test_verify_passes_with_expected_changes(sample_xlsx, config, tmp_path):
    out = tmp_path / "out_example.xlsx"
    process(sample_xlsx, out, config, ExampleAllTransform())
    report = verify(sample_xlsx, out, config)
    assert report.ok is True
    assert report.unexpected_changes == []
    assert len(report.expected_changes) > 0


def test_structural_attrs_unchanged(sample_xlsx, config, tmp_path):
    out = tmp_path / "out_example.xlsx"
    process(sample_xlsx, out, config, ExampleAllTransform())
    a = load_workbook(str(sample_xlsx))
    b = load_workbook(str(out))
    sa = a["Шаблон для поставщика"]
    sb = b["Шаблон для поставщика"]
    assert {str(r) for r in sa.merged_cells.ranges} == {str(r) for r in sb.merged_cells.ranges}
    wa = {k: v.width for k, v in sa.column_dimensions.items() if v.width}
    wb_ = {k: v.width for k, v in sb.column_dimensions.items() if v.width}
    assert wa == wb_
    dva = {str(dv.sqref) for dv in sa.data_validations.dataValidation}
    dvb = {str(dv.sqref) for dv in sb.data_validations.dataValidation}
    assert dva == dvb
    assert a.sheetnames == b.sheetnames
    a.close()
    b.close()
