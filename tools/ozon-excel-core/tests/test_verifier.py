"""Verifier proof works both ways: passes on legit edits, fails on frozen/header edits."""

from __future__ import annotations

from openpyxl import load_workbook

from ozon_excel_core import process, verify
from ozon_excel_core.transforms.example import ExampleAllTransform


def _process(sample_xlsx, config, tmp_path):
    out = tmp_path / "out.xlsx"
    process(sample_xlsx, out, config, ExampleAllTransform())
    return out


def test_legit_output_passes(sample_xlsx, config, tmp_path):
    out = _process(sample_xlsx, config, tmp_path)
    report = verify(sample_xlsx, out, config)
    assert report.ok is True
    assert report.unexpected_changes == []


def test_frozen_cell_tamper_fails(sample_xlsx, config, tmp_path):
    out = _process(sample_xlsx, config, tmp_path)
    # flip a frozen price cell (column G)
    wb = load_workbook(str(out))
    wb["Шаблон для поставщика"]["G4"] = 9999.99
    wb.save(str(out))
    wb.close()

    report = verify(sample_xlsx, out, config)
    assert report.ok is False
    bad = [d for d in report.unexpected_changes if d.coordinate == "G4"]
    assert bad, "expected G4 flagged as unexpected"
    assert bad[0].role == "frozen"


def test_header_cell_tamper_fails(sample_xlsx, config, tmp_path):
    out = _process(sample_xlsx, config, tmp_path)
    wb = load_workbook(str(out))
    wb["Шаблон для поставщика"]["B1"] = "ИЗМЕНЁННЫЙ ЗАГОЛОВОК"
    wb.save(str(out))
    wb.close()

    report = verify(sample_xlsx, out, config)
    assert report.ok is False
    bad = [d for d in report.unexpected_changes if d.coordinate == "B1"]
    assert bad
    assert bad[0].role == "header"


def test_verifier_fails_closed_on_ambiguous_mapping(sample_xlsx, tmp_path):
    """If an image keyword spec over-matches a frozen physical column, the
    mapper refuses — and the verifier must report a definitive FAIL rather than
    re-deriving the same broken allowed-set and certifying a leaked column as OK.
    This is what makes the verifier genuinely independent of mapper over-matches."""
    from dataclasses import replace as dc_replace

    from openpyxl import load_workbook

    from ozon_excel_core.config import parse_config

    # Build a config whose images_main keyword 'Главное фото' substring-matches a
    # frozen physical column we re-label on the sample.
    raw = {
        "version": 1,
        "sheets": [{"name": "Шаблон для поставщика"}],
        "header": {"scan_rows": 3, "match_row": 1, "data_start": 4},
        "match": {"mode": "contains"},
        "columns": {
            "title": {"keywords": {"ru": ["Название товара"]}, "required": True},
            "images_main": {"keywords": {"ru": ["Ссылка на главное фото", "Главное фото"]}},
        },
        "key_column": {"keywords": {"ru": ["Артикул"]}},
        "policy": {},
    }
    cfg = parse_config(raw)

    # Re-label a frozen column (G) to substring-contain the image keyword and
    # corrupt one of its data cells, simulating a buggy writer that leaked it.
    inp = tmp_path / "edge1d.xlsx"
    out = tmp_path / "edge1d_out.xlsx"
    wb = load_workbook(str(sample_xlsx))
    ws = wb["Шаблон для поставщика"]
    ws.cell(row=1, column=7, value="Главное фото для внутреннего архива")
    ws.cell(row=4, column=7, value="https://internal.supplier.test/private/orig.jpg")
    wb.save(str(inp))
    wb.close()

    wb = load_workbook(str(inp))
    wb["Шаблон для поставщика"].cell(
        row=4, column=7, value="https://cdn.example-ozon-mirror.test/private/orig.jpg"
    )
    wb.save(str(out))
    wb.close()

    report = verify(inp, out, cfg)
    assert report.ok is False
    assert any(d.field == "mapping" for d in report.unexpected_changes)


def test_cli_exit_codes_for_verify(sample_xlsx, config_path, tmp_path):
    from ozon_excel_core.cli import main

    out = tmp_path / "out.xlsx"
    rc = main(["process", "--in", str(sample_xlsx), "--out", str(out),
               "--config", config_path, "--transform", "example", "--quiet"])
    assert rc == 0
    rc = main(["verify", "--in", str(sample_xlsx), "--out", str(out), "--config", config_path])
    assert rc == 0

    # tamper -> exit 1
    wb = load_workbook(str(out))
    wb["Шаблон для поставщика"]["G4"] = 1
    wb.save(str(out))
    wb.close()
    rc = main(["verify", "--in", str(sample_xlsx), "--out", str(out), "--config", config_path])
    assert rc == 1
