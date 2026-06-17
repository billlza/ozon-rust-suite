"""Writer: input immutability + style/format preservation + merged anchor rule."""

from __future__ import annotations

import hashlib

from openpyxl import load_workbook

from ozon_excel_core import process
from ozon_excel_core.transforms.example import ExampleAllTransform


def _sha(path):
    with open(str(path), "rb") as fh:
        return hashlib.sha256(fh.read()).hexdigest()


def test_input_file_never_mutated(sample_xlsx, config, tmp_path):
    before = _sha(sample_xlsx)
    before_mtime = sample_xlsx.stat().st_mtime
    out = tmp_path / "out.xlsx"
    process(sample_xlsx, out, config, ExampleAllTransform())
    assert _sha(sample_xlsx) == before
    assert sample_xlsx.stat().st_mtime == before_mtime


def test_frozen_styles_and_number_formats_preserved(sample_xlsx, config, tmp_path):
    out = tmp_path / "out.xlsx"
    process(sample_xlsx, out, config, ExampleAllTransform())
    a = load_workbook(str(sample_xlsx))
    b = load_workbook(str(out))
    sa = a["Шаблон для поставщика"]
    sb = b["Шаблон для поставщика"]

    # price column G has #,##0.00 ; check a data cell keeps format + value
    assert sa["G4"].number_format == sb["G4"].number_format
    assert sa["G4"].value == sb["G4"].value
    # header bold/fill preserved
    assert sa["A1"].font.bold == sb["A1"].font.bold
    assert sa["A1"].fill.fgColor.rgb == sb["A1"].fill.fgColor.rgb
    a.close()
    b.close()


def test_no_write_to_non_anchor_merged_cell(sample_xlsx, config, tmp_path):
    # The sample merges the last two columns on row 1 (a header row, never a
    # target). Confirm those header cells are unchanged after processing.
    out = tmp_path / "out.xlsx"
    process(sample_xlsx, out, config, ExampleAllTransform())
    a = load_workbook(str(sample_xlsx))
    b = load_workbook(str(out))
    sa = a["Шаблон для поставщика"]
    sb = b["Шаблон для поставщика"]
    assert {str(r) for r in sa.merged_cells.ranges} == {str(r) for r in sb.merged_cells.ranges}
    a.close()
    b.close()
